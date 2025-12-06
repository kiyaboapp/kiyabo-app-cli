import os
import time
import random
import subprocess
import threading
import sys
import argparse
from typing import Optional, Tuple, List, Dict
import openpyxl
from openpyxl.styles import PatternFill
from rich.console import Console
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TimeRemainingColumn
from rich.prompt import Prompt, Confirm, IntPrompt
from rich.panel import Panel
from rich.table import Table
from rich.text import Text
from rich.layout import Layout
from rich import box
# Optional tkinter import - may not be available in all Python installations
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
    TKINTER_AVAILABLE = True
except ImportError:
    TKINTER_AVAILABLE = False
    tk = None
    filedialog = None
    messagebox = None
from tabulate import tabulate
from datetime import datetime

class SMSSender:
    def __init__(self, excel_path: str = "", delay_min: int = 5, delay_max: int = 10, 
                 show_window: bool = True, test_mode: bool = False):
        self.console = Console()
        self.should_cancel = False
        self.mpe_path = r"C:\Program Files (x86)\MyPhoneExplorer\MyPhoneExplorer.exe"
        self.excel_path = excel_path
        self.delay_min = delay_min
        self.delay_max = delay_max
        self.show_window = show_window
        self.test_mode = test_mode
        
        # Timing statistics
        self.timing_stats = {
            'min_time': float('inf'),
            'max_time': 0,
            'total_time': 0,
            'count': 0,
            'start_time': None,
            'end_time': None
        }
        
        self.stats = {
            'total': 0,
            'success': 0,
            'failed': 0,
            'skipped': 0,
            'cancelled': 0,
            'pending': 0,
            'already_sent': 0,
            'details': []
        }
        
        # Initialize Tkinter root (hidden) - only if available
        self.tkinter_available = TKINTER_AVAILABLE
        if self.tkinter_available:
            try:
                self.root = tk.Tk()
                self.root.withdraw()  # Hide the main window
            except Exception:
                self.tkinter_available = False
                self.root = None
        else:
            self.root = None
        
    def clear_screen(self):
        """Clear the console screen"""
        os.system('cls' if os.name == 'nt' else 'clear')
        
    def show_banner(self):
        """Display the ASCII art banner with colors"""
        banner = r'''
░██     ░██ ░██                      ░██                        ░██████   ░███     ░███   ░██████   
░██    ░██                           ░██                       ░██   ░██  ░████   ░████  ░██   ░██  
░██   ░██   ░██░██    ░██  ░██████   ░████████   ░███████     ░██         ░██░██ ░██░██ ░██         
░███████    ░██░██    ░██       ░██  ░██    ░██ ░██    ░██     ░████████  ░██ ░████ ░██  ░████████  
░██   ░██   ░██░██    ░██  ░███████  ░██    ░██ ░██    ░██            ░██ ░██  ░██  ░██         ░██ 
░██    ░██  ░██░██   ░███ ░██   ░██  ░███   ░██ ░██    ░██     ░██   ░██  ░██       ░██  ░██   ░██  
░██     ░██ ░██ ░█████░██  ░█████░██ ░██░█████   ░███████       ░██████   ░██       ░██   ░██████   
                      ░██                                                                           
                ░███████                                                                                                                                                                        
 '''

        self.console.print(banner, style="bold cyan")
        
    def show_header(self):
        """Display elegant header with banner"""
        self.clear_screen()
        self.show_banner()
        
        # Show current configuration in a colorful panel
        config_text = f"""[bold yellow]📁 Excel File:[/bold yellow] [cyan]{os.path.basename(self.excel_path) if self.excel_path else 'Not loaded'}[/cyan]
        [bold yellow]⏱️  Delay Range:[/bold yellow] [green]{self.delay_min}-{self.delay_max} seconds[/green]
        [bold yellow]🪟 Show Window:[/bold yellow] [magenta]{'Yes' if self.show_window else 'No'}[/magenta]
        [bold yellow]🧪 Test Mode:[/bold yellow] [blue]{'Yes' if self.test_mode else 'No'}[/blue]"""
        
        config_panel = Panel(
            config_text,
            title="[bold white]⚙️  Current Configuration[/bold white]",
            border_style="bright_blue",
            box=box.DOUBLE
        )
        self.console.print(config_panel)


        
    def show_menu(self):
        """Display colorful main menu with descriptions"""
        menu_items = [
            ("1", "📂 Load Excel File & Preview", "Load your SMS list and see a preview of messages"),
            ("2", "🔍 Check System Status", "Verify MyPhoneExplorer installation and connection"),
            ("3", "🔌 Connect to Phone", "Start MyPhoneExplorer and connect to your device"),
            ("4", "📤 Send SMS Messages", "Start sending messages from your Excel file"),
            ("5", "🧪 Test SMS Sending", "Send a test message to verify setup"),
            ("6", "⚙️  Adjust Settings", "Change delay times and window settings"),
            ("7", "📊 View Statistics", "See detailed statistics from previous sends"),
            ("0", "🚪 Exit Application", "Close the application")
        ]
        
        # Create a fancy menu
        self.console.print("╔═══════════════════════════════════════════════════════════════════════╗", style="bold bright_blue")
        self.console.print("║                          MAIN MENU                                    ║", style="bold bright_blue")
        self.console.print("╚═══════════════════════════════════════════════════════════════════════╝", style="bold bright_blue")
        self.console.print()
        
        for number, title, description in menu_items:
            # Colorful menu items
            self.console.print(f"  [{number}] ", style="bold bright_yellow", end="")
            self.console.print(f"{title}", style="bold white")
            # self.console.print(f"      [dim italic]{description}[/dim italic]")
            # self.console.print()
        
        self.console.print("─" * 75, style="dim bright_blue")
        self.console.print("\n💡 [bold yellow]Tip:[/bold yellow] [dim]Choose an option by entering its number[/dim]\n")
        self.console.print("👉 [bold bright_cyan]Enter your choice:[/bold bright_cyan] ", end="")
    

    
    def adjust_settings(self):
        """Allow user to adjust settings with better UX"""
        self.console.print("\n╔═══════════════════════════════════════════════════════════════════════╗", style="bold magenta")
        self.console.print("║                    ⚙️  ADJUST SETTINGS                                ║", style="bold magenta")
        self.console.print("╚═══════════════════════════════════════════════════════════════════════╝", style="bold magenta")
        
        # Show current settings
        current_settings = Panel(
            f"[yellow]Current Delay:[/yellow] [cyan]{self.delay_min}-{self.delay_max} seconds[/cyan]\n"
            f"[yellow]Window Visibility:[/yellow] [cyan]{'Shown' if self.show_window else 'Hidden'}[/cyan]",
            title="[bold white]Current Configuration[/bold white]",
            border_style="yellow"
        )
        self.console.print(current_settings)

        
        # Adjust delay settings
        self.console.print("⏱️  [bold cyan]Delay Settings[/bold cyan]", style="bold")
        self.console.print("   [dim]The delay between each SMS helps avoid being blocked by the carrier[/dim]\n")
        
        new_min = IntPrompt.ask(
            "   [yellow]Enter minimum delay in seconds[/yellow]",
            default=self.delay_min
        )
        new_max = IntPrompt.ask(
            "   [yellow]Enter maximum delay in seconds[/yellow]",
            default=self.delay_max
        )
        
        if new_min > new_max:
            self.console.print("\n❌ [bold red]Error:[/bold red] Minimum cannot be greater than maximum", style="red")
        else:
            self.delay_min = new_min
            self.delay_max = new_max
            self.console.print(f"\n✅ [bold green]Delay updated to {self.delay_min}-{self.delay_max} seconds[/bold green]")
        
        self.console.print()
        
        # Adjust window visibility
        self.console.print("🪟 [bold cyan]Window Visibility[/bold cyan]", style="bold")
        self.console.print("   [dim]Choose whether to show or hide MyPhoneExplorer window during sending[/dim]\n")
        
        self.show_window = Confirm.ask(
            "   [yellow]Show MyPhoneExplorer window?[/yellow]",
            default=self.show_window
        )
        
        window_status = "[green]shown[/green]" if self.show_window else "[yellow]hidden[/yellow]"
        self.console.print(f"\n✅ [bold green]Window will be {window_status}[/bold green]")
        
        # Show updated configuration in a panel
        self.console.print()
        updated_config = Panel(
            f"[yellow]Delay Range:[/yellow] [bright_green]{self.delay_min}-{self.delay_max} seconds[/bright_green]\n"
            f"[yellow]Show Window:[/yellow] [bright_green]{'Yes' if self.show_window else 'No'}[/bright_green]",
            title="[bold white]✨ Updated Configuration[/bold white]",
            border_style="bright_green",
            box=box.DOUBLE
        )
        self.console.print(updated_config)
    
    def test_sms_sending(self):
        """Test SMS sending with user-provided number and message"""
        self.console.print("\n╔═══════════════════════════════════════════════════════════════════════╗", style="bold bright_magenta")
        self.console.print("║                      🧪 TEST SMS SENDING                              ║", style="bold bright_magenta")
        self.console.print("╚═══════════════════════════════════════════════════════════════════════╝", style="bold bright_magenta")
        
        self.console.print("\n[dim]This will send a real SMS to verify your setup is working correctly[/dim]\n")
        
        if not self.is_sms_direct_installed():
            error_panel = Panel(
                "[bold red]MyPhoneExplorer is not installed on your system[/bold red]\n"
                "[yellow]Please install MyPhoneExplorer first to use this feature[/yellow]",
                title="❌ Error",
                border_style="red"
            )
            self.console.print(error_panel)
            return
        
        # Get phone number from user FIRST
        self.console.print("📞 [bold cyan]Enter Phone Number for Test SMS[/bold cyan]")
        self.console.print("   [dim]Example: 0712345678 or +255712345678[/dim]")
        phone_number = Prompt.ask("   [yellow]Phone number[/yellow]")
        
        if not phone_number or not phone_number.strip():
            error_panel = Panel(
                "[bold red]Phone number cannot be empty![/bold red]",
                title="❌ Error",
                border_style="red"
            )
            self.console.print(error_panel)
            return
        
        normalized_phone = self.normalize_phone(phone_number)
        if not normalized_phone or len(normalized_phone) < 12:
            error_panel = Panel(
                "[bold red]Invalid phone number format![/bold red]\n"
                f"[yellow]Could not normalize: {phone_number}[/yellow]",
                title="❌ Error",
                border_style="red"
            )
            self.console.print(error_panel)
            return
        
        self.console.print(f"   ✓ Normalized to: [bright_green]{normalized_phone}[/bright_green]\n")
        
        # Check if Excel file is loaded and has messages
        message = None
        first_message_available = False
        first_message_raw = ""
        first_student = ""
        
        if self.excel_path and os.path.exists(self.excel_path):
            try:
                workbook = openpyxl.load_workbook(self.excel_path)
                sheet = workbook.active
                # Check if there's a message in row 2, column 3 (SMS message column)
                if sheet.cell(row=2, column=3).value:
                    first_message_available = True
                    first_message_raw = str(sheet.cell(row=2, column=3).value or "").strip()
                    first_student = str(sheet.cell(row=2, column=2).value or "").strip()
                workbook.close()
            except Exception as e:
                self.console.print(f"[dim yellow]⚠️  Could not read Excel file: {str(e)}[/dim yellow]\n")
        
        # Let user choose message source
        if first_message_available and first_message_raw:
            self.console.print("💬 [bold cyan]Choose Message Source[/bold cyan]")
            self.console.print("   [1] ✏️  [white]Write a new message[/white]")
            self.console.print(f"   [2] 📋 [white]Use first message from Excel table[/white]")
            if first_student:
                self.console.print(f"       [dim](From record: {first_student})[/dim]")
            self.console.print()
            
            msg_choice = Prompt.ask("   [yellow]Choose option[/yellow]", choices=["1", "2"], default="2")
            
            if msg_choice == '2':
                message = first_message_raw
                # Show preview of the message that will be used - FULL MESSAGE, NO TRUNCATION
                self.console.print()
                # Escape Rich markup to prevent rendering issues
                preview_display = str(first_message_raw).replace('[', '\\[').replace(']', '\\]')
                preview_panel = Panel(
                    f"[white]{preview_display}[/white]",
                    title=f"[bold bright_green]📋 First Message from Excel (Student: {first_student or 'N/A'})[/bold bright_green]",
                    border_style="bright_green",
                    box=box.ROUNDED,
                    expand=False
                )
                self.console.print(preview_panel)
                self.console.print()
            else:
                self.console.print("\n💬 [bold cyan]Enter Test Message[/bold cyan]")
                self.console.print("   [dim]Enter the message you want to send[/dim]")
                message = Prompt.ask("   [yellow]Message[/yellow]")
        else:
            # No Excel file loaded or no message available, just ask for message
            if not self.excel_path:
                self.console.print("[dim]ℹ️  No Excel file loaded. Enter message manually.[/dim]\n")
            elif not first_message_available:
                self.console.print("[dim]ℹ️  Excel file loaded but no message found in first record.[/dim]\n")
            
            self.console.print("💬 [bold cyan]Enter Test Message[/bold cyan]")
            self.console.print("   [dim]Enter the message you want to send[/dim]")
            message = Prompt.ask("   [yellow]Message[/yellow]")
        
        # Validate message - must exist and not be empty after sanitization
        if not message or not message.strip():
            error_panel = Panel(
                "[bold red]Message cannot be empty![/bold red]",
                title="❌ Error",
                border_style="red"
            )
            self.console.print(error_panel)
            return
        
        # Show what will be sent (raw and sanitized) - FULL MESSAGE, NO TRUNCATION
        sanitized_message = self.sanitize_sms(message)
        sanitized_message = sanitized_message.strip()
        
        # Validate sanitized message is not empty
        if not sanitized_message:
            error_panel = Panel(
                "[bold red]Message is empty after sanitization![/bold red]",
                title="❌ Error",
                border_style="red"
            )
            self.console.print(error_panel)
            return
        self.console.print()
        self.console.print("📋 [bold bright_cyan]Message Preview (Full Message):[/bold bright_cyan]")
        self.console.print("─" * 75, style="bright_cyan")
        
        # Escape Rich markup to prevent rendering issues, but show FULL message
        raw_display = str(message).replace('[', '\\[').replace(']', '\\]')
        raw_preview = Panel(
            f"[white]{raw_display}[/white]",
            title="[bold yellow]Original Message (Before Sanitization)[/bold yellow]",
            border_style="yellow",
            box=box.ROUNDED,
            expand=False
        )
        self.console.print(raw_preview)
        self.console.print()
        
        if message != sanitized_message:
            # Escape Rich markup in sanitized message too, show FULL message
            sanitized_display = str(sanitized_message).replace('[', '\\[').replace(']', '\\]')
            sanitized_preview = Panel(
                f"[bright_green]{sanitized_display}[/bright_green]",
                title="[bold bright_green]Sanitized Message (What Will Be Sent)[/bold bright_green]",
                border_style="bright_green",
                box=box.ROUNDED,
                expand=False
            )
            self.console.print(sanitized_preview)
            self.console.print()
            self.console.print("[dim]ℹ️  Note: Newlines have been converted to %n for SMS compatibility[/dim]")
            self.console.print()
        
        # Show final confirmation panel
        confirm_panel = Panel(
            f"[bold yellow]To:[/bold yellow] [bright_cyan]{normalized_phone}[/bright_cyan]\n"
            f"[bold yellow]Message Length:[/bold yellow] [white]{len(message)} characters[/white]\n"
            f"[bold yellow]Sanitized Length:[/bold yellow] [white]{len(sanitized_message)} characters[/white]",
            title="[bold white]📋 Test SMS Details - Ready to Send[/bold white]",
            border_style="yellow",
            box=box.DOUBLE
        )
        self.console.print(confirm_panel)
        self.console.print()
        
        if not Confirm.ask("[bold yellow]Send this test SMS?[/bold yellow]", default=True):
            self.console.print("\n❌ [yellow]Test cancelled[/yellow]")
            return
        
        # Send test SMS with progress indicator and 60-second timeout
        self.console.print()
        self.console.print("[dim]⏱️  Timeout set to 60 seconds - will stop automatically if error occurs[/dim]")
        self.console.print()
        with self.console.status("[bold yellow]🚀 Sending test SMS (60s timeout)...[/bold yellow]", spinner="dots"):
            success, result_message = self.try_send_text_sms(normalized_phone, message, timeout=60)
        
        self.console.print()
        if success:
            success_panel = Panel(
                f"[bold bright_green]✅ Test SMS sent successfully![/bold bright_green]\n"
                f"[dim]{result_message}[/dim]",
                title="Success",
                border_style="bright_green",
                box=box.DOUBLE
            )
            self.console.print(success_panel)
        else:
            error_panel = Panel(
                f"[bold red]❌ Test SMS failed![/bold red]\n"
                f"[yellow]{result_message}[/yellow]",
                title="Error",
                border_style="red"
            )
            self.console.print(error_panel)
    
    def browse_excel_file(self) -> str:
        """Open file dialog to browse for Excel file"""
        if not self.tkinter_available or filedialog is None:
            # Fallback to manual path input if tkinter is not available
            self.console.print("[yellow]⚠️  File dialog not available (tkinter not installed)[/yellow]")
            self.console.print("[dim]Please enter the file path manually[/dim]\n")
            file_path = Prompt.ask("[yellow]📁 Enter full path to Excel file[/yellow]")
            return file_path.strip('"\'')
        
        try:
            self.console.print("[dim]Opening file browser...[/dim]")
            file_path = filedialog.askopenfilename(
                title="Select Excel File",
                filetypes=[
                    ("Excel files", "*.xlsx *.xls"),
                    ("All files", "*.*")
                ]
            )
            return file_path
        except Exception as e:
            self.console.print(f"❌ [red]File dialog error: {str(e)}[/red]")
            # Fallback to manual input
            self.console.print("[yellow]Falling back to manual path input...[/yellow]\n")
            file_path = Prompt.ask("[yellow]📁 Enter full path to Excel file[/yellow]")
            return file_path.strip('"\'')
    
    def is_sms_direct_installed(self) -> bool:
        """Check if MyPhoneExplorer is installed"""
        return os.path.exists(self.mpe_path)
    
    def is_app_running(self, app_name: str = "MyPhoneExplorer.exe") -> bool:
        """Check if MyPhoneExplorer is running"""
        try:
            result = subprocess.run(
                ['tasklist', '/FI', f'IMAGENAME eq {app_name}'],
                capture_output=True, text=True, creationflags=subprocess.CREATE_NO_WINDOW
            )
            return app_name.lower() in result.stdout.lower()
        except:
            return False
    
    def connect_myphone_explorer(self) -> bool:
        """Initialize connection to MyPhoneExplorer"""
        self.console.print("\n╔═══════════════════════════════════════════════════════════════════════╗", style="bold bright_cyan")
        self.console.print("║                   🔌 CONNECT TO PHONE                                 ║", style="bold bright_cyan")
        self.console.print("╚═══════════════════════════════════════════════════════════════════════╝", style="bold bright_cyan")
        self.console.print()
        
        if not self.is_sms_direct_installed():
            error_panel = Panel(
                "[bold red]MyPhoneExplorer is not installed[/bold red]\n\n"
                "[yellow]Please download and install MyPhoneExplorer from:[/yellow]\n"
                "[cyan]https://www.fjsoft.at/en/[/cyan]",
                title="❌ Installation Required",
                border_style="red"
            )
            self.console.print(error_panel)
            return False
        
        try:
            with self.console.status("[bold yellow]🔌 Connecting to MyPhoneExplorer...[/bold yellow]", spinner="dots"):
                creation_flags = 0 if self.show_window else subprocess.CREATE_NO_WINDOW
                args = "action=connect"
                command = f'"{self.mpe_path}" {args}'
                
                process = subprocess.run(
                    command,
                    shell=True,
                    capture_output=True,
                    text=True,
                    timeout=30,
                    creationflags=creation_flags
                )
            
            self.console.print()
            if process.returncode == 0:
                success_panel = Panel(
                    "[bold bright_green]✅ MyPhoneExplorer connected successfully![/bold bright_green]\n"
                    "[dim]Your phone is now connected and ready to send messages[/dim]",
                    title="Success",
                    border_style="bright_green",
                    box=box.DOUBLE
                )
                self.console.print(success_panel)
                return True
            else:
                warning_panel = Panel(
                    "[bold yellow]⚠️  MyPhoneExplorer opened[/bold yellow]\n"
                    "[dim]Please manually connect your phone in the MyPhoneExplorer window[/dim]",
                    title="Manual Setup Required",
                    border_style="yellow"
                )
                self.console.print(warning_panel)
                return True
                
        except subprocess.TimeoutExpired:
            success_panel = Panel(
                "[bold green]✅ MyPhoneExplorer started[/bold green]\n"
                "[dim]The application is now running[/dim]",
                title="Success",
                border_style="green"
            )
            self.console.print(success_panel)
            return True
        except Exception as e:
            error_panel = Panel(
                f"[bold red]❌ Connection failed[/bold red]\n"
                f"[yellow]{str(e)}[/yellow]",
                title="Error",
                border_style="red"
            )
            self.console.print(error_panel)
            return False
    
    def sanitize_sms(self, message: str) -> str:
        """Sanitize SMS message - replace newlines with %n for SMS sending"""
        if not message:
            return ""
        # Convert to string if not already
        text = str(message)
        # Handle various newline representations:
        # 1. Standard newlines and carriage returns
        text = text.replace('\r\n', '%n')  # Windows line endings first
        text = text.replace('\r', '%n')     # Carriage returns
        text = text.replace('\n', '%n')     # Line feeds
        # 2. Handle literal string representations that might come from Excel
        text = text.replace('_x000D_', '%n')  # Excel carriage return representation
        text = text.replace('_x000A_', '%n')  # Excel line feed representation
        # Remove any duplicate %n sequences (optional - keeps message cleaner)
        while '%n%n' in text:
            text = text.replace('%n%n', '%n')
        return text
    
    def normalize_phone(self, phone: str) -> str:
        """Normalize phone number to Tanzanian format"""
        if not phone:
            return ""
            
        phone = str(phone).strip().replace(" ", "")
        cleaned_phone = ''.join(c for c in phone if c.isdigit() or c == '+')
        
        if cleaned_phone.startswith('255') and len(cleaned_phone) == 12:
            return "+" + cleaned_phone
        elif len(cleaned_phone) == 9:
            return "+255" + cleaned_phone
        elif len(cleaned_phone) == 10 and cleaned_phone.startswith('0'):
            return "+255" + cleaned_phone[1:]
        elif cleaned_phone.startswith('+'):
            return cleaned_phone
        else:
            return "+255" + cleaned_phone[-9:]
    
    def is_valid_sms(self, phone_raw: str, message_raw: str) -> Tuple[bool, Optional[str], Optional[str]]:
        """Check if SMS has both valid phone number and valid message
        
        Returns:
            Tuple of (is_valid: bool, clean_phone: Optional[str], text_sms: Optional[str])
            clean_phone and text_sms are None if invalid, otherwise contain normalized values
        """
        # Validate phone number - must normalize and be at least 12 characters
        clean_phone = None
        if phone_raw and phone_raw.strip():
            normalized = self.normalize_phone(phone_raw)
            if normalized and len(normalized) >= 12:  # Valid phone format
                clean_phone = normalized
        
        # Validate message - must exist and not be empty after sanitization
        text_sms = None
        if message_raw:
            sanitized = self.sanitize_sms(message_raw)
            sanitized = sanitized.strip()
            if sanitized:  # Not empty after sanitization
                text_sms = sanitized
        
        # Both must be valid
        is_valid = clean_phone is not None and text_sms is not None
        return (is_valid, clean_phone, text_sms)
    
    def count_valid_sms(self, sheet, start_row: int, end_row: int) -> int:
        """Count SMS records with both valid phone number and valid message"""
        valid_count = 0
        for i in range(start_row, end_row + 1):
            phone_raw = str(sheet.cell(row=i, column=4).value or "")
            message_raw = str(sheet.cell(row=i, column=3).value or "")
            
            is_valid, _, _ = self.is_valid_sms(phone_raw, message_raw)
            if is_valid:
                valid_count += 1
        return valid_count
    
    def load_excel_preview(self) -> bool:
        """Load Excel file and show preview"""
        self.console.print("\n╔═══════════════════════════════════════════════════════════════════════╗", style="bold bright_green")
        self.console.print("║                📂 LOAD EXCEL FILE & PREVIEW                           ║", style="bold bright_green")
        self.console.print("╚═══════════════════════════════════════════════════════════════════════╝", style="bold bright_green")
        self.console.print()
        
        if not self.excel_path:
            self.console.print("[bold cyan]How would you like to load your Excel file?[/bold cyan]\n")
            self.console.print("  [1] 🗔  [white]Browse with file dialog[/white] [dim](Recommended)[/dim]")
            self.console.print("  [2] ⌨️  [white]Enter path manually[/white]")
            self.console.print()
            
            choice = Prompt.ask("[yellow]Choose option[/yellow]", choices=["1", "2"], default="1")
            
            if choice == '1':
                self.console.print()
                file_path = self.browse_excel_file()
                if file_path:
                    self.excel_path = file_path
                    self.console.print(f"✅ [green]Selected: {os.path.basename(self.excel_path)}[/green]")
                else:
                    self.console.print("❌ [red]No file selected[/red]")
                    return False
            else:
                self.console.print()
                path_input = Prompt.ask("[yellow]📁 Enter full path to Excel file[/yellow]")
                self.excel_path = path_input.strip('"\'')
        
        self.excel_path = os.path.abspath(self.excel_path)
        
        if not os.path.exists(self.excel_path):
            error_panel = Panel(
                f"[bold red]File not found![/bold red]\n"
                f"[yellow]Path: {self.excel_path}[/yellow]",
                title="❌ Error",
                border_style="red"
            )
            self.console.print(error_panel)
            self.excel_path = ""
            return False
        
        try:
            with self.console.status("[bold yellow]📖 Loading Excel file...[/bold yellow]", spinner="dots"):
                # Create backup
                backup_path = self.excel_path.replace('.xlsx', f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
                import shutil
                shutil.copy2(self.excel_path, backup_path)
                
                workbook = openpyxl.load_workbook(self.excel_path)
                sheet = workbook.active
                
                # Find data range
                last_row = 1
                for row in range(2, sheet.max_row + 1):
                    if sheet.cell(row=row, column=4).value:
                        last_row = row
                    else:
                        break
            
            self.console.print()
            
            if last_row <= 1:
                error_panel = Panel(
                    "[bold red]No data found in Excel file[/bold red]\n"
                    "[yellow]Please make sure your Excel file has data starting from row 2[/yellow]",
                    title="❌ Empty File",
                    border_style="red"
                )
                self.console.print(error_panel)
                workbook.close()
                return False
            
            total_rows = last_row - 1
            valid_sms_count = self.count_valid_sms(sheet, 2, last_row)
            
            # Success header
            success_panel = Panel(
                f"[bold bright_green]✅ Excel file loaded successfully![/bold bright_green]\n\n"
                f"[yellow]File:[/yellow] [cyan]{os.path.basename(self.excel_path)}[/cyan]\n"
                f"[yellow]Total Records:[/yellow] [bright_cyan]{total_rows}[/bright_cyan]\n"
                f"[yellow]Valid SMS (Phone + Message):[/yellow] [bright_green]{valid_sms_count}[/bright_green]\n"
                f"[yellow]Backup Created:[/yellow] [dim]{os.path.basename(backup_path)}[/dim]",
                title="📊 File Information",
                border_style="bright_green",
                box=box.DOUBLE
            )
            self.console.print(success_panel)
            self.console.print()
            
            # Get header row (row 1) to determine column names
            headers = []
            max_col = sheet.max_column
            for col in range(1, max_col + 1):
                header_value = str(sheet.cell(row=1, column=col).value or f"Column {col}").strip()
                headers.append(header_value)
            
            # Show Excel data preview in table format (5 records)
            self.console.print("📋 [bold bright_cyan]Excel Data Preview (First 5 Records):[/bold bright_cyan]")
            self.console.print("═" * 75, style="bright_cyan")
            self.console.print()
            
            # Create table with all columns - show lines between rows for better readability
            preview_table = Table(
                title="Raw Excel Data",
                box=box.DOUBLE,
                border_style="bright_blue",
                show_header=True,
                show_lines=True  # Add dividing lines between rows
            )
            
            # Add columns based on headers - no truncation, show full content
            for header in headers:
                preview_table.add_column(header, style="white", overflow="fold", no_wrap=False)
            
            # Add 5 rows of data - show full content without truncation
            for i in range(2, min(7, last_row + 1)):
                row_data = []
                for col in range(1, max_col + 1):
                    cell_value = str(sheet.cell(row=i, column=col).value or "").strip()
                    # Show full content - no truncation
                    row_data.append(cell_value)
                preview_table.add_row(*row_data)
            
            self.console.print(preview_table)
            self.console.print()
            
            # Show SMS message for first record (raw and escaped)
            if last_row >= 2:
                first_student = str(sheet.cell(row=2, column=2).value or "").strip()
                first_phone = str(sheet.cell(row=2, column=4).value or "")
                first_message_raw = str(sheet.cell(row=2, column=3).value or "")
                first_status = str(sheet.cell(row=2, column=5).value or "").strip()
                
                normalized_phone = self.normalize_phone(first_phone)
                
                self.console.print("💬 [bold bright_magenta]First Record SMS Message Preview:[/bold bright_magenta]")
                self.console.print("═" * 75, style="bright_magenta")
                self.console.print()
                
                # Show first record details
                first_record_panel = Panel(
                    f"[bold yellow]Student:[/bold yellow] [white]{first_student}[/white]\n"
                    f"[bold yellow]Phone:[/bold yellow] [cyan]{first_phone}[/cyan] → [bright_green]{normalized_phone}[/bright_green]\n"
                    f"[bold yellow]Status:[/bold yellow] [magenta]{first_status if first_status else 'PENDING'}[/magenta]",
                    title="[bold white]Record #1 Details[/bold white]",
                    border_style="bright_cyan",
                    box=box.ROUNDED
                )
                self.console.print(first_record_panel)
                self.console.print()
                
                # Show raw message - escape Rich markup to prevent rendering issues
                # Replace any Rich markup characters that might be in the message
                raw_display = str(first_message_raw).replace('[', '\\[').replace(']', '\\]')
                raw_message_panel = Panel(
                    f"[white]{raw_display}[/white]",
                    title="[bold yellow]📝 Raw SMS Message (Before Escaping)[/bold yellow]",
                    border_style="yellow",
                    box=box.ROUNDED,
                    expand=False
                )
                self.console.print(raw_message_panel)
                self.console.print()
                
                # Show escaped message - this is what will be sent
                escaped_message = self.sanitize_sms(first_message_raw)
                # Escape Rich markup in escaped message too
                escaped_display = str(escaped_message).replace('[', '\\[').replace(']', '\\]')
                escaped_message_panel = Panel(
                    f"[bright_green]{escaped_display}[/bright_green]",
                    title="[bold bright_green]✨ Escaped SMS Message (After Sanitization)[/bold bright_green]",
                    border_style="bright_green",
                    box=box.ROUNDED,
                    expand=False
                )
                self.console.print(escaped_message_panel)
                self.console.print()
                
                # Show what was changed
                if first_message_raw != escaped_message:
                    changes_info = Panel(
                        "[dim]Note: Newlines (\\n) and carriage returns (\\r) have been replaced with %n[/dim]",
                        border_style="dim",
                        box=box.SIMPLE
                    )
                    self.console.print(changes_info)
                    self.console.print()
                else:
                    no_changes_info = Panel(
                        "[dim]No non-printable characters found - message is ready to send[/dim]",
                        border_style="dim",
                        box=box.SIMPLE
                    )
                    self.console.print(no_changes_info)
                    self.console.print()
            
            # Status summary - ONLY count status for valid SMS (both valid phone and valid message)
            status_counts = {}
            for i in range(2, last_row + 1):
                phone_raw = str(sheet.cell(row=i, column=4).value or "")
                message_raw = str(sheet.cell(row=i, column=3).value or "")
                
                # Only count status if both phone and message are valid
                is_valid, _, _ = self.is_valid_sms(phone_raw, message_raw)
                if is_valid:
                    status = str(sheet.cell(row=i, column=5).value or "").strip().upper()
                    if status:
                        status_counts[status] = status_counts.get(status, 0) + 1
                    else:
                        status_counts['PENDING'] = status_counts.get('PENDING', 0) + 1
            
            # Comprehensive Statistics Summary
            self.console.print()
            self.console.print("📊 [bold bright_cyan]Comprehensive Statistics Summary:[/bold bright_cyan]")
            self.console.print("═" * 75, style="bright_cyan")
            self.console.print()
            
            # Calculate detailed statistics
            valid_phones = 0
            invalid_phones = 0
            valid_messages = 0
            empty_messages = 0
            
            for i in range(2, last_row + 1):
                phone_raw = str(sheet.cell(row=i, column=4).value or "")
                message_raw = str(sheet.cell(row=i, column=3).value or "")
                
                if phone_raw and phone_raw.strip():
                    normalized = self.normalize_phone(phone_raw)
                    if normalized and len(normalized) >= 12:  # Valid phone format
                        valid_phones += 1
                    else:
                        invalid_phones += 1
                else:
                    invalid_phones += 1
                
                if message_raw and message_raw.strip():
                    valid_messages += 1
                else:
                    empty_messages += 1
            
            # Show file statistics panel
            file_stats_panel = Panel(
                f"[bold yellow]Total Records:[/bold yellow] [bright_cyan]{total_rows}[/bright_cyan]\n"
                f"[bold yellow]Valid SMS (Phone + Message):[/bold yellow] [bright_green]{valid_sms_count}[/bright_green]\n"
                f"[bold yellow]Valid Phone Numbers:[/bold yellow] [bright_green]{valid_phones}[/bright_green]\n"
                f"[bold yellow]Invalid Phone Numbers:[/bold yellow] [red]{invalid_phones}[/red]\n"
                f"[bold yellow]Valid Messages:[/bold yellow] [bright_green]{valid_messages}[/bright_green]\n"
                f"[bold yellow]Empty Messages:[/bold yellow] [yellow]{empty_messages}[/yellow]",
                title="[bold white]📈 File Statistics[/bold white]",
                border_style="bright_cyan",
                box=box.ROUNDED
            )
            self.console.print(file_stats_panel)
            self.console.print()
            
            # Status summary table
            if status_counts:
                # Create colored status table - print only once
                table = Table(title="📈 Current Status Summary", box=box.DOUBLE, border_style="bright_yellow")
                table.add_column("Status", style="bold", justify="left")
                table.add_column("Count", style="bright_cyan", justify="right")
                table.add_column("Percentage", style="bright_green", justify="right")
                
                for status, count in sorted(status_counts.items()):
                    percentage = (count / valid_sms_count) * 100 if valid_sms_count > 0 else 0
                    # Color code based on status
                    if status in ["OK", "SENT"]:
                        status_display = f"[bright_green]✅ {status}[/bright_green]"
                    elif status == "PENDING":
                        status_display = f"[yellow]⏳ {status}[/yellow]"
                    elif status == "FAILED":
                        status_display = f"[red]❌ {status}[/red]"
                    else:
                        status_display = f"[blue]{status}[/blue]"
                    
                    table.add_row(status_display, str(count), f"{percentage:.1f}%")
                
                self.console.print(table)
                self.console.print()  # Add spacing after table
            
            workbook.close()
            return True
            
        except Exception as e:
            error_panel = Panel(
                f"[bold red]Error loading Excel file[/bold red]\n\n"
                f"[yellow]{str(e)}[/yellow]",
                title="❌ Error",
                border_style="red"
            )
            self.console.print(error_panel)
            self.excel_path = ""
            return False
    
    def try_send_text_sms(self, phone_number: str, text_sms: str, timeout: int = 30) -> Tuple[bool, str]:
        """Send SMS using MyPhoneExplorer and wait for completion
        
        Args:
            phone_number: Phone number to send SMS to
            text_sms: SMS message text
            timeout: Timeout in seconds (default 30, use 60 for test SMS)
        
        Returns:
            Tuple of (success: bool, message: str)
        """
        if not self.is_sms_direct_installed():
            return False, "MyPhoneExplorer not installed"
        
        start_time = time.time()
        process = None
        timeout_occurred = threading.Event()
        
        def kill_on_timeout():
            """Forcefully kill process if timeout is exceeded"""
            time.sleep(timeout)
            if process and process.poll() is None:
                timeout_occurred.set()
                # Try to terminate gracefully first
                try:
                    process.terminate()
                    process.wait(timeout=2)
                except:
                    pass
                # Force kill if still running
                if process.poll() is None:
                    try:
                        process.kill()
                        process.wait()
                    except:
                        pass
        
        try:
            sanitized_message = self.sanitize_sms(text_sms)
            args = f'action=sendmessage savetosent=1 number={phone_number} Text="{sanitized_message}"'
            command = f'"{self.mpe_path}" {args}'
            creation_flags = 0 if self.show_window else subprocess.CREATE_NO_WINDOW
            
            # Use Popen for better timeout control
            process = subprocess.Popen(
                command,
                shell=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                creationflags=creation_flags
            )
            
            # Start timeout monitor thread
            timeout_thread = threading.Thread(target=kill_on_timeout, daemon=True)
            timeout_thread.start()
            
            # Wait for process to complete
            stdout, stderr = process.communicate()
            returncode = process.returncode
            
            # Check if timeout occurred
            if timeout_occurred.is_set():
                elapsed = time.time() - start_time
                return False, f"Timeout after {timeout}s (elapsed: {elapsed:.2f}s) - SMS sending took too long or encountered an error"
            
            execution_time = time.time() - start_time
            
            # Update timing statistics
            self.timing_stats['min_time'] = min(self.timing_stats['min_time'], execution_time)
            self.timing_stats['max_time'] = max(self.timing_stats['max_time'], execution_time)
            self.timing_stats['total_time'] += execution_time
            self.timing_stats['count'] += 1
            
            # Wait random delay after sending (only for bulk sending, not test)
            if timeout == 30:  # Bulk sending
                delay = random.randint(self.delay_min, self.delay_max)
                time.sleep(delay)
                delay_info = f" (waited {delay}s)"
            else:  # Test sending
                delay_info = ""
            
            if returncode == 0:
                return True, f"Sent in {execution_time:.2f}s{delay_info}"
            else:
                error_msg = stderr.strip() if stderr else stdout.strip() if stdout else "Unknown error"
                return False, f"Failed in {execution_time:.2f}s: {error_msg}{delay_info}"
                
        except Exception as e:
            # Make sure to kill process if it's still running
            if process and process.poll() is None:
                try:
                    process.terminate()
                    process.wait(timeout=2)
                except:
                    try:
                        process.kill()
                        process.wait()
                    except:
                        pass
            elapsed = time.time() - start_time
            if timeout_occurred.is_set() or elapsed >= timeout:
                return False, f"Timeout after {timeout}s (elapsed: {elapsed:.2f}s) - SMS sending took too long or encountered an error"
            return False, f"Error: {str(e)}"
    
    def send_sms_messages(self):
        """Main function to send SMS messages with enhanced confirmations"""
        self.console.print("\n╔═══════════════════════════════════════════════════════════════════════╗", style="bold bright_magenta")
        self.console.print("║                    📤 SEND SMS MESSAGES                               ║", style="bold bright_magenta")
        self.console.print("╚═══════════════════════════════════════════════════════════════════════╝", style="bold bright_magenta")
        self.console.print()
        
        # Pre-flight checks
        if not self.excel_path:
            error_panel = Panel(
                "[bold red]No Excel file loaded![/bold red]\n\n"
                "[yellow]Please load an Excel file first using option 1[/yellow]",
                title="❌ Error",
                border_style="red"
            )
            self.console.print(error_panel)
            return
        
        if not self.is_sms_direct_installed():
            error_panel = Panel(
                "[bold red]MyPhoneExplorer is not installed![/bold red]\n\n"
                "[yellow]Please install MyPhoneExplorer to continue[/yellow]",
                title="❌ Error",
                border_style="red"
            )
            self.console.print(error_panel)
            return
        
        # Check if MyPhoneExplorer is running
        if not self.is_app_running():
            warning_panel = Panel(
                "[bold yellow]⚠️  MyPhoneExplorer is not running![/bold yellow]\n\n"
                "[dim]The application needs to be connected to your phone to send messages[/dim]",
                title="Connection Required",
                border_style="yellow"
            )
            self.console.print(warning_panel)
            self.console.print()
            
            if Confirm.ask("[bold cyan]Would you like to start and connect it now?[/bold cyan]", default=True):
                if not self.connect_myphone_explorer():
                    error_panel = Panel(
                        "[bold red]Cannot proceed without MyPhoneExplorer connection[/bold red]",
                        title="❌ Error",
                        border_style="red"
                    )
                    self.console.print(error_panel)
                    return
            else:
                self.console.print("\n❌ [red]Cannot send SMS without MyPhoneExplorer[/red]")
                return
        
        try:
            workbook = openpyxl.load_workbook(self.excel_path)
            sheet = workbook.active
            
            # Find last row
            last_row = 1
            for row in range(2, sheet.max_row + 1):
                if sheet.cell(row=row, column=4).value:
                    last_row = row
                else:
                    break
            
            if last_row <= 1:
                error_panel = Panel(
                    "[bold red]No data found in Excel file[/bold red]",
                    title="❌ Error",
                    border_style="red"
                )
                self.console.print(error_panel)
                workbook.close()
                return
            
            total_rows = last_row - 1
            valid_sms_count = self.count_valid_sms(sheet, 2, last_row)
            
            # Check for already sent messages (only count valid SMS)
            already_sent_count = 0
            pending_count = 0
            for i in range(2, last_row + 1):
                phone_raw = str(sheet.cell(row=i, column=4).value or "")
                message_raw = str(sheet.cell(row=i, column=3).value or "")
                
                # Only count if both phone and message are valid
                is_valid, _, _ = self.is_valid_sms(phone_raw, message_raw)
                if is_valid:
                    status = str(sheet.cell(row=i, column=5).value or "").strip().upper()
                    if status in ["OK", "SENT"]:
                        already_sent_count += 1
                    else:
                        pending_count += 1
            
            # Show summary
            summary_panel = Panel(
                f"[bold yellow]Total Records:[/bold yellow] [bright_cyan]{total_rows}[/bright_cyan]\n"
                f"[bold yellow]Valid SMS (Phone + Message):[/bold yellow] [bright_green]{valid_sms_count}[/bright_green]\n"
                f"[bold yellow]Already Sent:[/bold yellow] [bright_green]{already_sent_count}[/bright_green]\n"
                f"[bold yellow]Pending:[/bold yellow] [bright_yellow]{pending_count}[/bright_yellow]",
                title="📊 Summary",
                border_style="bright_cyan",
                box=box.DOUBLE
            )
            self.console.print(summary_panel)
            self.console.print()
            
            resend_all = False
            if already_sent_count > 0:
                self.console.print(f"[bold yellow]ℹ️  Found {already_sent_count} messages already marked as SENT/OK[/bold yellow]\n")
                resend_all = Confirm.ask(
                    "[bold cyan]Do you want to resend ALL messages (including already sent)?[/bold cyan]",
                    default=False
                )
                if not resend_all:
                    self.console.print("✅ [green]Will skip already sent messages[/green]\n")
                else:
                    self.console.print("⚠️  [yellow]Will resend ALL messages[/yellow]\n")
            
            # Delay settings confirmation
            self.console.print("[bold cyan]📋 Current Delay Settings:[/bold cyan]")
            self.console.print(f"   Delay between messages: [bright_yellow]{self.delay_min}-{self.delay_max} seconds[/bright_yellow]\n")
            
            if Confirm.ask("[bold cyan]Do you want to adjust the delay settings?[/bold cyan]", default=False):
                self.console.print()
                new_min = IntPrompt.ask(
                    "   [yellow]Minimum delay (seconds)[/yellow]",
                    default=self.delay_min
                )
                new_max = IntPrompt.ask(
                    "   [yellow]Maximum delay (seconds)[/yellow]",
                    default=self.delay_max
                )
                
                if new_min <= new_max:
                    self.delay_min = new_min
                    self.delay_max = new_max
                    self.console.print(f"\n✅ [green]Delay updated to {self.delay_min}-{self.delay_max} seconds[/green]\n")
                else:
                    self.console.print("\n⚠️  [yellow]Invalid range, keeping current settings[/yellow]\n")
            
            # Final confirmation with all details
            messages_to_send = valid_sms_count if resend_all else pending_count
            estimated_min_time = messages_to_send * self.delay_min / 60
            estimated_max_time = messages_to_send * self.delay_max / 60
            
            final_confirmation = Panel(
                f"[bold white]Ready to Send SMS Messages[/bold white]\n\n"
                f"[yellow]Messages to Send:[/yellow] [bright_cyan]{messages_to_send}[/bright_cyan]\n"
                f"[yellow]Delay Range:[/yellow] [bright_green]{self.delay_min}-{self.delay_max} seconds[/bright_green]\n"
                f"[yellow]Estimated Time:[/yellow] [bright_magenta]{estimated_min_time:.1f}-{estimated_max_time:.1f} minutes[/bright_magenta]\n"
                f"[yellow]Window:[/yellow] [blue]{'Shown' if self.show_window else 'Hidden'}[/blue]\n\n"
                f"[dim]You can press 'q' + Enter anytime to cancel[/dim]",
                title="🚀 Final Confirmation",
                border_style="bright_yellow",
                box=box.DOUBLE
            )
            self.console.print(final_confirmation)
            self.console.print()
            
            if not Confirm.ask("[bold bright_green]START SENDING SMS NOW?[/bold bright_green]", default=False):
                self.console.print("\n❌ [yellow]Sending cancelled by user[/yellow]")
                workbook.close()
                return
            
            # Reset timing stats for this session
            self.timing_stats = {
                'min_time': float('inf'),
                'max_time': 0,
                'total_time': 0,
                'count': 0,
                'start_time': datetime.now(),
                'end_time': None
            }
            
            self.stats = {
                'total': valid_sms_count,
                'success': 0,
                'failed': 0,
                'skipped': 0,
                'cancelled': 0,
                'pending': 0,
                'already_sent': already_sent_count,
                'details': []
            }
            
            self.console.print()
            start_panel = Panel(
                "[bold bright_green]🚀 Starting SMS sending process...[/bold bright_green]\n"
                "[yellow]⏸️  Press 'q' + Enter to cancel at any time[/yellow]",
                border_style="bright_green",
                box=box.DOUBLE
            )
            self.console.print(start_panel)
            self.console.print()
            
            # Start cancellation listener
            self.should_cancel = False
            cancel_thread = threading.Thread(target=self._listen_for_cancel)
            cancel_thread.daemon = True
            cancel_thread.start()
            
            # Process with progress bar
            with Progress(
                SpinnerColumn(style="bright_cyan"),
                TextColumn("[progress.description]{task.description}"),
                BarColumn(complete_style="bright_green", finished_style="bright_green"),
                TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
                TextColumn("•"),
                TimeRemainingColumn(),
                console=self.console
            ) as progress:
                
                task = progress.add_task("[bold bright_cyan]Sending SMS...", total=valid_sms_count)
                processed_count = 0
                valid_processed_count = 0
                invalid_count = 0
                
                for i in range(2, last_row + 1):
                    if self.should_cancel:
                        remaining = valid_sms_count - valid_processed_count
                        self.stats['cancelled'] = remaining
                        break
                    
                    student_name = str(sheet.cell(row=i, column=2).value or "").strip()
                    phone_raw = str(sheet.cell(row=i, column=4).value or "")
                    message_raw = str(sheet.cell(row=i, column=3).value or "")
                    current_status = str(sheet.cell(row=i, column=5).value or "").strip().upper()
                    
                    status_cell = sheet.cell(row=i, column=5)
                    detail_cell = sheet.cell(row=i, column=6)
                    
                    # Validate SMS - must have both valid phone and valid message
                    is_valid, clean_phone, text_sms = self.is_valid_sms(phone_raw, message_raw)
                    
                    # Only process valid SMS (both phone and message valid)
                    if is_valid:
                        valid_processed_count += 1
                        
                        # Skip already sent messages unless resend_all is True
                        if not resend_all and current_status in ["OK", "SENT"]:
                            self.stats['skipped'] += 1
                            progress.console.print(f"[dim]⏭️  {student_name}: Already sent - skipping[/dim]")
                            progress.update(task, advance=1)
                            continue
                        
                        progress.console.print(f"[bold yellow]📤 Sending to {student_name}...[/bold yellow]")
                        
                        send_result, status_message = self.try_send_text_sms(clean_phone, text_sms)
                        
                        if send_result:
                            status_cell.value = "OK"
                            status_cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                            self.stats['success'] += 1
                            progress.console.print(f"[bright_green]✅ {student_name}: {status_message}[/bright_green]")
                        else:
                            status_cell.value = "PENDING"
                            status_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                            self.stats['failed'] += 1
                            progress.console.print(f"[red]❌ {student_name}: {status_message}[/red]")
                        
                        detail_cell.value = status_message
                        
                        # Store details for statistics
                        self.stats['details'].append({
                            'student': student_name,
                            'phone': clean_phone,
                            'status': 'OK' if send_result else 'PENDING',
                            'message': status_message
                        })
                        
                        progress.update(task, advance=1)
                    else:
                        # Invalid SMS - mark but don't count in progress (silently skip to avoid spam)
                        status_cell.value = "INVALID"
                        status_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        detail_cell.value = "Invalid phone or message"
                        invalid_count += 1
                        # Only show invalid message summary at the end, not for each one
                    
                    processed_count += 1
                
                # Show summary of invalid records if any
                if invalid_count > 0:
                    progress.console.print(f"[dim]⚠️  Skipped {invalid_count} invalid records (missing phone or message)[/dim]")
                
                # Update end time
                self.timing_stats['end_time'] = datetime.now()
                
                # Save workbook
                self.console.print()
                if not self.should_cancel:
                    with self.console.status("[bold yellow]💾 Saving results to Excel...[/bold yellow]", spinner="dots"):
                        try:
                            workbook.save(self.excel_path)
                        except Exception as e:
                            # Try to save with a different name
                            new_path = self.excel_path.replace('.xlsx', f'_updated_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
                            workbook.save(new_path)
                            self.console.print(f"[yellow]💾 Saved as backup: {os.path.basename(new_path)}[/yellow]")
                    
                    self.console.print("[bright_green]✅ Excel file saved with status updates[/bright_green]")
                else:
                    self.console.print("[yellow]⚠️  Changes not saved (cancelled)[/yellow]")
            
            workbook.close()
            
            # Show final statistics
            self.console.print()
            self.show_statistics()
            
        except Exception as e:
            error_panel = Panel(
                f"[bold red]An error occurred[/bold red]\n\n"
                f"[yellow]{str(e)}[/yellow]",
                title="❌ Error",
                border_style="red"
            )
            self.console.print(error_panel)
    
    def show_statistics(self):
        """Display detailed statistics in colorful format"""
        stats = self.stats
        timing = self.timing_stats
        
        self.console.print()
        self.console.print("╔═══════════════════════════════════════════════════════════════════════╗", style="bold bright_cyan")
        self.console.print("║                   📊 SMS SENDING STATISTICS                           ║", style="bold bright_cyan")
        self.console.print("╚═══════════════════════════════════════════════════════════════════════╝", style="bold bright_cyan")
        self.console.print()
        
        # Main statistics table with colors
        table = Table(title="📈 Results Summary", box=box.DOUBLE, border_style="bright_yellow")
        table.add_column("Category", style="bold white", justify="left")
        table.add_column("Count", style="bright_cyan", justify="right")
        table.add_column("Percentage", style="bright_green", justify="right")
        
        table.add_row("[bold]Total Messages[/bold]", str(stats['total']), "-")
        
        if stats['success'] > 0:
            success_pct = (stats['success'] / stats['total']) * 100
            table.add_row("[bright_green]✅ Successfully Sent[/bright_green]", str(stats['success']), f"{success_pct:.1f}%")
        
        if stats['failed'] > 0:
            failed_pct = (stats['failed'] / stats['total']) * 100
            table.add_row("[red]❌ Failed[/red]", str(stats['failed']), f"{failed_pct:.1f}%")
        
        if stats['skipped'] > 0:
            skipped_pct = (stats['skipped'] / stats['total']) * 100
            table.add_row("[yellow]⚠️  Skipped (Invalid)[/yellow]", str(stats['skipped']), f"{skipped_pct:.1f}%")
        
        if stats['already_sent'] > 0:
            already_pct = (stats['already_sent'] / stats['total']) * 100
            table.add_row("[dim]⏭️  Already Sent[/dim]", str(stats['already_sent']), f"{already_pct:.1f}%")
        
        if stats['cancelled'] > 0:
            cancelled_pct = (stats['cancelled'] / stats['total']) * 100
            table.add_row("[bright_magenta]🚫 Cancelled[/bright_magenta]", str(stats['cancelled']), f"{cancelled_pct:.1f}%")
        
        self.console.print(table)
        self.console.print()
        
        # Timing statistics
        if timing['count'] > 0:
            avg_time = timing['total_time'] / timing['count']
            
            timing_table = Table(title="⏱️  Timing Statistics", box=box.ROUNDED, border_style="bright_magenta")
            timing_table.add_column("Metric", style="bold yellow", justify="left")
            timing_table.add_column("Value", style="bright_cyan", justify="right")
            
            timing_table.add_row("Total Execution Time", f"{timing['total_time']:.2f}s")
            timing_table.add_row("Number of SMS Sent", str(timing['count']))
            timing_table.add_row("Fastest SMS", f"[bright_green]{timing['min_time']:.2f}s[/bright_green]")
            timing_table.add_row("Slowest SMS", f"[bright_yellow]{timing['max_time']:.2f}s[/bright_yellow]")
            timing_table.add_row("Average Time per SMS", f"[bright_cyan]{avg_time:.2f}s[/bright_cyan]")
            
            if timing['start_time'] and timing['end_time']:
                duration = timing['end_time'] - timing['start_time']
                timing_table.add_row("Total Session Duration", str(duration).split('.')[0])
            
            self.console.print(timing_table)
            self.console.print()
        
        # Recent activity table
        if stats['details']:
            recent_count = min(8, len(stats['details']))
            recent_table = Table(
                title=f"📋 Recent Activity (Showing {recent_count} of {len(stats['details'])} processed)",
                box=box.SIMPLE,
                border_style="bright_blue"
            )
            recent_table.add_column("Status", justify="center")
            recent_table.add_column("Student", style="white")
            recent_table.add_column("Phone", style="cyan")
            recent_table.add_column("Result", style="dim")
            
            for detail in stats['details'][:8]:
                status_icon = "[bright_green]✅[/bright_green]" if detail['status'] == 'OK' else "[red]❌[/red]"
                recent_table.add_row(
                    status_icon,
                    detail['student'],
                    detail['phone'],
                    detail['message']
                )
            
            self.console.print(recent_table)
    
    def show_app_status(self):
        """Show MyPhoneExplorer installation and connection status"""
        self.console.print("\n╔═══════════════════════════════════════════════════════════════════════╗", style="bold bright_blue")
        self.console.print("║                    🔍 SYSTEM STATUS CHECK                             ║", style="bold bright_blue")
        self.console.print("╚═══════════════════════════════════════════════════════════════════════╝", style="bold bright_blue")
        self.console.print()
        
        # Create status table
        table = Table(box=box.DOUBLE, border_style="bright_cyan")
        table.add_column("Component", style="bold yellow", justify="left")
        table.add_column("Status", justify="left")
        table.add_column("Details", style="dim", justify="left")
        
        # Installation status
        installed = self.is_sms_direct_installed()
        if installed:
            table.add_row(
                "MyPhoneExplorer",
                "[bright_green]✅ Installed[/bright_green]",
                self.mpe_path
            )
        else:
            table.add_row(
                "MyPhoneExplorer",
                "[red]❌ Not Installed[/red]",
                "Please install MyPhoneExplorer"
            )
        
        # Running status
        running = self.is_app_running()
        if running:
            table.add_row(
                "Connection",
                "[bright_green]✅ Running[/bright_green]",
                "MyPhoneExplorer is active"
            )
        else:
            table.add_row(
                "Connection",
                "[yellow]⚠️  Not Running[/yellow]",
                "Start MyPhoneExplorer (option 3)"
            )
        
        # Excel status
        excel_loaded = self.excel_path and os.path.exists(self.excel_path)
        if excel_loaded:
            try:
                workbook = openpyxl.load_workbook(self.excel_path)
                sheet = workbook.active
                last_row = 1
                for row in range(2, sheet.max_row + 1):
                    if sheet.cell(row=row, column=4).value:
                        last_row = row
                valid_sms_count = self.count_valid_sms(sheet, 2, last_row)
                workbook.close()
                table.add_row(
                    "Excel File",
                    "[bright_green]✅ Loaded[/bright_green]",
                    f"{valid_sms_count} valid SMS ready"
                )
            except:
                table.add_row(
                    "Excel File",
                    "[yellow]⚠️  File Issues[/yellow]",
                    "Cannot read file properly"
                )
        else:
            table.add_row(
                "Excel File",
                "[red]❌ Not Loaded[/red]",
                "Load file (option 1)"
            )
        
        self.console.print(table)
        self.console.print()
        
        # Overall readiness
        if installed and running and excel_loaded:
            ready_panel = Panel(
                "[bold bright_green]✅ System is ready to send SMS![/bold bright_green]\n"
                "[dim]You can proceed to option 4 to start sending messages[/dim]",
                title="System Ready",
                border_style="bright_green",
                box=box.DOUBLE
            )
            self.console.print(ready_panel)
        else:
            issues = []
            if not installed:
                issues.append("Install MyPhoneExplorer")
            if not running:
                issues.append("Start MyPhoneExplorer (option 3)")
            if not excel_loaded:
                issues.append("Load Excel file (option 1)")
            
            warning_panel = Panel(
                "[bold yellow]⚠️  System not ready[/bold yellow]\n\n"
                "[white]Please complete these steps:[/white]\n" +
                "\n".join([f"  • {issue}" for issue in issues]),
                title="Action Required",
                border_style="yellow"
            )
            self.console.print(warning_panel)
    
    def _listen_for_cancel(self):
        """Listen for cancellation input"""
        while not self.should_cancel:
            try:
                user_input = input().strip().lower()
                if user_input == 'q':
                    self.should_cancel = True
                    self.console.print("\n[bold yellow]🛑 Cancellation requested... stopping after current message[/bold yellow]")
                    break
            except:
                break
    
    def run(self):
        """Main application loop"""
        while True:
            self.show_header()
            self.show_menu()
            
            try:
                choice = input().strip()
                
                if choice == '1':
                    self.load_excel_preview()
                elif choice == '2':
                    self.show_app_status()
                elif choice == '3':
                    self.connect_myphone_explorer()
                elif choice == '4':
                    self.send_sms_messages()
                elif choice == '5':
                    self.test_sms_sending()
                elif choice == '6':
                    self.adjust_settings()
                elif choice == '7':
                    if self.stats['total'] > 0:
                        self.show_statistics()
                    else:
                        info_panel = Panel(
                            "[yellow]No statistics available yet[/yellow]\n"
                            "[dim]Statistics will appear after you send messages[/dim]",
                            title="📊 Statistics",
                            border_style="yellow"
                        )
                        self.console.print(info_panel)
                elif choice == '0':
                    goodbye_panel = Panel(
                        "[bold bright_green]Thank you for using BlueZA App![/bold bright_green]\n"
                        "[dim]SMS Sending System by BlueZA[/dim]",
                        title="👋 Goodbye",
                        border_style="bright_green",
                        box=box.DOUBLE
                    )
                    self.console.print("\n")
                    self.console.print(goodbye_panel)
                    sys.exit(0)
                else:
                    error_panel = Panel(
                        f"[bold red]'{choice}' is not a valid option[/bold red]\n"
                        "[yellow]Please enter a number between 0-7[/yellow]",
                        title="❌ Invalid Choice",
                        border_style="red"
                    )
                    self.console.print("\n")
                    self.console.print(error_panel)
                    
                # Wait for user to continue
                if choice != '0':
                    self.console.print("\n" + "─" * 75, style="dim bright_blue")
                    self.console.print("[bold bright_cyan]Press Enter to return to main menu...[/bold bright_cyan]", end="")
                    input()
                    
            except KeyboardInterrupt:
                self.console.print("\n\n")
                interrupt_panel = Panel(
                    "[bold yellow]Application interrupted by user[/bold yellow]\n"
                    "[dim]Goodbye![/dim]",
                    title="👋 Exit",
                    border_style="yellow"
                )
                self.console.print(interrupt_panel)
                sys.exit(0)
            except Exception as e:
                error_panel = Panel(
                    f"[bold red]An unexpected error occurred[/bold red]\n\n"
                    f"[yellow]{str(e)}[/yellow]",
                    title="❌ Error",
                    border_style="red"
                )
                self.console.print("\n")
                self.console.print(error_panel)
                self.console.print("\n[dim]Press Enter to continue...[/dim]", end="")
                input()

def parse_arguments():
    """Parse command line arguments"""
    parser = argparse.ArgumentParser(description='SMS Sending System using MyPhoneExplorer')
    parser.add_argument('--excel', '-e', type=str, help='Path to Excel file')
    parser.add_argument('--delay-min', '-min', type=int, default=5, help='Minimum delay between SMS (seconds)')
    parser.add_argument('--delay-max', '-max', type=int, default=10, help='Maximum delay between SMS (seconds)')
    parser.add_argument('--show-window', '-w', action='store_true', help='Show MyPhoneExplorer window')
    parser.add_argument('--test', '-t', action='store_true', help='Enable test mode')
    
    return parser.parse_args()

def main():
    """Main function with command line argument support"""
    args = parse_arguments()
    
    app = SMSSender(
        excel_path=args.excel,
        delay_min=args.delay_min,
        delay_max=args.delay_max,
        show_window=args.show_window,
        test_mode=args.test
    )
    
    app.run()

if __name__ == "__main__":
    main()