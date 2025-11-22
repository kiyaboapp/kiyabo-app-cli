#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import xml.etree.ElementTree as ET
from xml.dom import minidom
import subprocess
import re
import sys
import time
import random
import textwrap
from pathlib import Path
from datetime import datetime
from tabulate import tabulate
from colorama import init, Fore, Back, Style
from tqdm import tqdm

init(autoreset=True)

class BulkSMSSender:
    """Clean, professional SMS sender with direct message and batch XML modes"""
    
    def __init__(self,
                 excel_path: str,
                 mpe_exe_path: str,
                 sheet_name: str = None,
                 mode: str = "direct",  # "direct" or "batch"
                 batch_size: int = 1,
                 min_sleep: int = 8,
                 max_sleep: int = 12):
        """
        Initialize SMS Sender
        
        Args:
            excel_path: Path to Excel file with contacts
            mpe_exe_path: Path to MyPhoneExplorer.exe
            sheet_name: Excel sheet name (default: first sheet)
            mode: "direct" (one-by-one) or "batch" (XML files)
            batch_size: Messages per batch (for batch mode)
            min_sleep: Minimum seconds between sends
            max_sleep: Maximum seconds between sends
        """
        self.excel_path = Path(excel_path)
        self.mpe_exe_path = Path(mpe_exe_path)
        self.sheet_name = sheet_name
        self.mode = mode.lower()
        self.batch_size = batch_size
        self.min_sleep = min_sleep
        self.max_sleep = max_sleep
        
        # Excel objects
        self.wb = None
        self.ws = None
        self.header_map = {}
        
        # Output directory for batch XMLs
        self.output_dir = None
        
        # Styling
        self.ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.ok_font = Font(color="006100", bold=True)
        self.fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.fail_font = Font(color="9C0006", bold=True)
        self.pending_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        self.pending_font = Font(color="806000")
        
        # Stats
        self.total_sent = 0
        self.total_success = 0
        self.total_failed = 0
        
    def print_banner(self, text: str, color=Fore.CYAN):
        """Print a styled banner"""
        width = 100
        print(f"\n{color}{Style.BRIGHT}{'═' * width}")
        print(f"{text.center(width)}")
        print(f"{'═' * width}{Style.RESET_ALL}\n")
    
    def print_section(self, text: str):
        """Print a section header"""
        print(f"\n{Fore.YELLOW}{Style.BRIGHT}▶ {text}{Style.RESET_ALL}")
        print(f"{Fore.YELLOW}{'─' * 100}{Style.RESET_ALL}\n")
    
    def load_excel(self):
        """Load and validate Excel file"""
        self.print_banner("📱 BULK SMS SENDER", Fore.CYAN)
        
        if not self.excel_path.exists():
            sys.exit(f"{Fore.RED}✘ Excel file not found: {self.excel_path}{Style.RESET_ALL}")
        
        if not self.mpe_exe_path.exists():
            sys.exit(f"{Fore.RED}✘ MyPhoneExplorer not found: {self.mpe_exe_path}{Style.RESET_ALL}")
        
        print(f"{Fore.GREEN}✓ Loading Excel: {self.excel_path.name}{Style.RESET_ALL}")
        self.wb = openpyxl.load_workbook(self.excel_path)
        self.ws = self.wb[self.sheet_name] if self.sheet_name else self.wb[self.wb.sheetnames[0]]
        
        # Map headers
        for cell in self.ws[1]:
            header = str(cell.value or "").strip().upper()
            if header:
                self.header_map[header] = cell.column
        
        # Validate required columns
        required = ["PHONE NUMBER", "SMS TO BE SENT", "FULL NAME"]
        missing = [col for col in required if col not in self.header_map]
        if missing:
            sys.exit(f"{Fore.RED}✘ Missing required columns: {', '.join(missing)}{Style.RESET_ALL}")
        
        # Add STATUS column if missing
        if "STATUS" not in self.header_map:
            col = self.ws.max_column + 1
            letter = get_column_letter(col)
            self.ws[f"{letter}1"] = "STATUS"
            self.ws[f"{letter}1"].font = Font(bold=True)
            self.ws[f"{letter}1"].alignment = Alignment(horizontal="center")
            self.header_map["STATUS"] = col
            
            # Set all rows to PENDING
            for row in range(2, self.ws.max_row + 1):
                cell = self.ws.cell(row, col)
                cell.value = "PENDING"
                cell.fill = self.pending_fill
                cell.font = self.pending_font
                cell.alignment = Alignment(horizontal="center")
        
        print(f"{Fore.GREEN}✓ Sheet: {self.ws.title} ({self.ws.max_row - 1} records){Style.RESET_ALL}")
        print(f"{Fore.GREEN}✓ Mode: {self.mode.upper()}{Style.RESET_ALL}")
        print(f"{Fore.GREEN}✓ Delay: {self.min_sleep}-{self.max_sleep} seconds{Style.RESET_ALL}")
        
    def normalize_phone(self, raw):
        """Normalize phone number to 255XXXXXXXXX format"""
        if not raw:
            return None
        
        digits = re.sub(r'\D', '', str(raw))
        if not digits:
            return None
        
        # Already in correct format
        if digits.startswith("255") and len(digits) == 12:
            return digits
        
        # Remove leading 0 and add 255
        if digits.startswith("0") and len(digits) == 10:
            return "255" + digits[1:]
        
        # Add 255 prefix
        if len(digits) == 9:
            return "255" + digits
        
        return None
    
    def sanitize_sms(self, text: str) -> str:
        """Sanitize SMS text for command line - replace newlines with %n"""
        if not text:
            return ""
        # Replace newlines with %n for MyPhoneExplorer
        text = str(text).replace('\n', '%n').replace('\r', '')
        return text
    
    def show_preview(self):
        """Show first 20 records as preview - FULL MESSAGE TEXT, NO TRUNCATION"""
        self.print_section("📋 PREVIEW - First 20 Records (FULL TEXT)")
        
        table_data = []
        for row in range(2, min(22, self.ws.max_row + 1)):
            name = str(self.ws.cell(row, self.header_map["FULL NAME"]).value or "")
            phone_raw = self.ws.cell(row, self.header_map["PHONE NUMBER"]).value
            phone = self.normalize_phone(phone_raw)
            sms = str(self.ws.cell(row, self.header_map["SMS TO BE SENT"]).value or "")
            status = str(self.ws.cell(row, self.header_map["STATUS"]).value or "PENDING")
            
            phone_display = phone if phone else f"{Fore.RED}INVALID{Style.RESET_ALL}"
            
            # Wrap message text at reasonable width for display
            wrapped_sms = textwrap.fill(sms, width=80, break_long_words=False, break_on_hyphens=False)
            
            table_data.append([
                row - 1,
                name,
                phone_display,
                wrapped_sms,
                status
            ])
        
        # Use simple_grid for clean table with headers at top
        print(tabulate(
            table_data,
            headers=["#", "Name", "Phone", "Message (Full Text)", "Status"],
            tablefmt="grid"
        ))
        print()
    
    def get_pending_records(self):
        """Get all records that need to be sent (not OK status)"""
        self.print_section("🔍 Scanning Records")
        
        to_send = []
        invalid = []
        
        for row in range(2, self.ws.max_row + 1):
            status_cell = self.ws.cell(row, self.header_map["STATUS"])
            status = str(status_cell.value or "").strip().upper()
            
            # Skip already sent messages
            if status == "OK":
                continue
            
            # Get record data
            name = str(self.ws.cell(row, self.header_map["FULL NAME"]).value or "Unknown")
            phone_raw = self.ws.cell(row, self.header_map["PHONE NUMBER"]).value
            phone = self.normalize_phone(phone_raw)
            sms = str(self.ws.cell(row, self.header_map["SMS TO BE SENT"]).value or "")
            
            # Validate phone
            if not phone:
                invalid.append([row - 1, name[:30], str(phone_raw or "(empty)")])
                continue
            
            # Reset to PENDING if status is something else
            if status != "PENDING":
                status_cell.value = "PENDING"
                status_cell.fill = self.pending_fill
                status_cell.font = self.pending_font
            
            to_send.append({
                "row": row,
                "name": name,
                "phone": phone,
                "sms": sms,
                "status_cell": status_cell
            })
        
        # Show invalid numbers in MULTI-COLUMN format to save space
        if invalid:
            print(f"{Fore.RED}{Style.BRIGHT}✘ Invalid Phone Numbers: {len(invalid)}{Style.RESET_ALL}\n")
            
            # Split into 3 columns for compact display
            columns = 3
            multi_col_data = []
            
            for i in range(0, len(invalid), columns):
                row_data = []
                for j in range(columns):
                    if i + j < len(invalid):
                        rec = invalid[i + j]
                        row_data.extend([rec[0], rec[1], rec[2]])  # #, Name, Phone
                    else:
                        row_data.extend(["", "", ""])  # Empty cells
                multi_col_data.append(row_data)
            
            # Headers for 3 columns
            headers = []
            for col in range(columns):
                headers.extend([f"#{col+1}", f"Name{col+1}", f"Phone{col+1}"])
            
            print(tabulate(multi_col_data, headers=headers, tablefmt="fancy_grid"))
            print()
        
        print(f"{Fore.GREEN}{Style.BRIGHT}✓ Ready to send: {len(to_send)} messages{Style.RESET_ALL}\n")
        return to_send
    
    def send_direct(self, record: dict):
        """Send single message directly via MyPhoneExplorer command with 2-minute timeout"""
        phone = record["phone"]
        text = self.sanitize_sms(record["sms"])
        
        # Build command - direct shell execution
        cmd = f'"{self.mpe_exe_path}" action=sendmessage savetosent=1 number={phone} text="{text}" simcard=2'
        
        print(f"{Fore.CYAN}→ Sending to {record['name']} ({phone})...{Style.RESET_ALL}", end=" ")
        
        try:
            # Execute command with 120 second (2 minute) timeout
            result = subprocess.run(
                cmd,
                shell=True,
                capture_output=True,
                timeout=120  # 2 minutes timeout
            )
            
            # Check shell return code - if 0, command executed successfully
            if result.returncode == 0:
                # Command executed = message sent successfully
                record["status_cell"].value = "OK"
                record["status_cell"].fill = self.ok_fill
                record["status_cell"].font = self.ok_font
                print(f"{Fore.GREEN}{Style.BRIGHT}✓ OK{Style.RESET_ALL}")
                self.total_success += 1
                return True
            else:
                # Non-zero return code = execution problem
                record["status_cell"].value = f"FAILED (code {result.returncode})"
                record["status_cell"].fill = self.fail_fill
                record["status_cell"].font = self.fail_font
                print(f"{Fore.RED}{Style.BRIGHT}✘ FAILED (code {result.returncode}){Style.RESET_ALL}")
                self.total_failed += 1
                return False
                
        except subprocess.TimeoutExpired:
            # Timeout after 2 minutes = ABORT WHOLE PROCESS
            record["status_cell"].value = "TIMEOUT - ABORTED"
            record["status_cell"].fill = self.fail_fill
            record["status_cell"].font = self.fail_font
            print(f"\n\n{Fore.RED}{Back.YELLOW}{Style.BRIGHT}{'='*100}{Style.RESET_ALL}")
            print(f"{Fore.RED}{Back.YELLOW}{Style.BRIGHT}⚠ CRITICAL TIMEOUT - PROCESS ABORTED{Style.RESET_ALL}")
            print(f"{Fore.RED}{Back.YELLOW}{Style.BRIGHT}{'='*100}{Style.RESET_ALL}\n")
            print(f"{Fore.RED}✘ No response after 2 minutes for {record['name']} ({phone}){Style.RESET_ALL}")
            print(f"{Fore.YELLOW}⚠ MyPhoneExplorer may not be responding or phone not connected{Style.RESET_ALL}")
            print(f"{Fore.YELLOW}⚠ Please check:{Style.RESET_ALL}")
            print(f"   1. MyPhoneExplorer is running")
            print(f"   2. Phone is connected via USB/WiFi")
            print(f"   3. Phone has network signal")
            print(f"   4. You have SMS credits remaining")
            print(f"\n{Fore.CYAN}Excel file will be saved with current progress...{Style.RESET_ALL}\n")
            self.total_failed += 1
            raise  # Re-raise to stop the whole process
            
        except Exception as e:
            # Any other exception = problem
            record["status_cell"].value = "ERROR"
            record["status_cell"].fill = self.fail_fill
            record["status_cell"].font = self.fail_font
            print(f"{Fore.RED}{Style.BRIGHT}✘ ERROR: {str(e)}{Style.RESET_ALL}")
            self.total_failed += 1
            return False
    
    def create_batch_xml(self, batch: list, batch_num: int) -> Path:
        """Create XML file for batch sending"""
        root = ET.Element("batch")
        
        for record in batch:
            message = ET.SubElement(root, "message")
            recipient = ET.SubElement(message, "recipient", status="PENDING")
            recipient.text = record["phone"]
            text = ET.SubElement(message, "text")
            text.text = record["sms"]
        
        # Pretty print XML
        xml_str = ET.tostring(root, encoding="unicode")
        dom = minidom.parseString(xml_str)
        pretty_xml = dom.toprettyxml(indent="  ")
        
        # Save to file
        xml_path = self.output_dir / f"batch_{batch_num:03d}.xml"
        xml_path.write_text(pretty_xml, encoding="utf-8")
        
        return xml_path
    
    def send_batch_xml(self, xml_path: Path, batch: list):
        """Send batch via XML file and check results"""
        cmd = f'"{self.mpe_exe_path}" action=sendmessage batchfile="{xml_path.resolve()}"'
        
        print(f"\n{Fore.CYAN}{Style.BRIGHT}▶ Sending batch: {xml_path.name} ({len(batch)} messages){Style.RESET_ALL}")
        print(f"{Fore.WHITE}Command: {cmd}{Style.RESET_ALL}\n")
        
        try:
            # Launch MyPhoneExplorer
            subprocess.Popen(cmd, shell=True, creationflags=subprocess.CREATE_NEW_CONSOLE)
            
            # Wait for processing
            wait_time = max(30, len(batch) * 10)
            print(f"{Fore.YELLOW}⏳ Waiting {wait_time} seconds for MPE to process...{Style.RESET_ALL}")
            
            for i in range(wait_time, 0, -1):
                print(f"   {i:02d}s remaining...", end="\r")
                time.sleep(1)
            
            print(f"\n{Fore.GREEN}✓ Processing complete. Checking results...{Style.RESET_ALL}\n")
            
            # Parse results from XML
            self.parse_batch_results(xml_path, batch)
            
        except Exception as e:
            print(f"{Fore.RED}✘ Error sending batch: {e}{Style.RESET_ALL}")
            for record in batch:
                record["status_cell"].value = "ERROR"
                record["status_cell"].fill = self.fail_fill
                record["status_cell"].font = self.fail_font
                self.total_failed += 1
    
    def parse_batch_results(self, xml_path: Path, batch: list):
        """Parse XML results and update Excel"""
        try:
            tree = ET.parse(xml_path)
            root = tree.getroot()
            
            results_table = []
            
            for i, message in enumerate(root.findall("message")):
                if i >= len(batch):
                    break
                
                record = batch[i]
                recipient = message.find("recipient")
                status = recipient.get("status", "UNKNOWN") if recipient is not None else "UNKNOWN"
                
                if status == "OK":
                    record["status_cell"].value = "OK"
                    record["status_cell"].fill = self.ok_fill
                    record["status_cell"].font = self.ok_font
                    status_display = f"{Fore.GREEN}{Style.BRIGHT}OK{Style.RESET_ALL}"
                    self.total_success += 1
                else:
                    record["status_cell"].value = status
                    record["status_cell"].fill = self.fail_fill
                    record["status_cell"].font = self.fail_font
                    status_display = f"{Fore.RED}{Style.BRIGHT}{status}{Style.RESET_ALL}"
                    self.total_failed += 1
                
                results_table.append([
                    i + 1,
                    record["name"][:30],
                    record["phone"],
                    status_display
                ])
            
            print(tabulate(
                results_table,
                headers=["#", "Name", "Phone", "Status"],
                tablefmt="fancy_grid"
            ))
            
        except ET.ParseError:
            print(f"{Fore.RED}✘ XML not modified - MyPhoneExplorer did not process the file{Style.RESET_ALL}")
            for record in batch:
                record["status_cell"].value = "FAILED"
                record["status_cell"].fill = self.fail_fill
                record["status_cell"].font = self.fail_font
                self.total_failed += 1
    
    def run(self):
        """Main execution method"""
        # Load and validate
        self.load_excel()
        self.show_preview()
        
        # Get records to send
        records = self.get_pending_records()
        
        if not records:
            print(f"{Fore.YELLOW}ℹ No messages to send. All done!{Style.RESET_ALL}")
            return
        
        # Confirmation
        self.print_section("📤 Ready to Send")
        print(f"Mode: {Fore.CYAN}{self.mode.upper()}{Style.RESET_ALL}")
        print(f"Messages: {Fore.CYAN}{len(records)}{Style.RESET_ALL}")
        print(f"Delay: {Fore.CYAN}{self.min_sleep}-{self.max_sleep} seconds{Style.RESET_ALL}")
        
        if self.mode == "batch":
            print(f"Batch size: {Fore.CYAN}{self.batch_size}{Style.RESET_ALL}")
            self.output_dir = Path(f"sms_batches_{datetime.now():%Y%m%d_%H%M%S}")
            self.output_dir.mkdir(exist_ok=True)
            print(f"XML folder: {Fore.CYAN}{self.output_dir}{Style.RESET_ALL}")
        
        print()
        
        # Start sending
        start_time = datetime.now()
        self.print_banner("🚀 SENDING MESSAGES", Fore.GREEN)
        
        if self.mode == "direct":
            # Direct one-by-one sending with progress bar
            with tqdm(total=len(records), 
                     desc=f"{Fore.GREEN}Sending SMS{Style.RESET_ALL}",
                     bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}]",
                     colour="green") as pbar:
                
                for i, record in enumerate(records, 1):
                    pbar.set_description(f"{Fore.GREEN}Sending to {record['name'][:30]}{Style.RESET_ALL}")
                    
                    self.send_direct(record)
                    self.total_sent += 1
                    
                    # Update progress bar
                    pbar.update(1)
                    
                    # Save after each message
                    self.wb.save(self.excel_path)
                    
                    # Wait between messages
                    if i < len(records):
                        wait = random.randint(self.min_sleep, self.max_sleep)
                        pbar.set_description(f"{Fore.YELLOW}Waiting {wait}s...{Style.RESET_ALL}")
                        time.sleep(wait)
        
        else:
            # Batch XML sending with progress bar
            batch_num = 1
            i = 0
            total_batches = (len(records) + self.batch_size - 1) // self.batch_size
            
            with tqdm(total=total_batches,
                     desc=f"{Fore.GREEN}Sending Batches{Style.RESET_ALL}",
                     bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}]",
                     colour="green") as pbar:
                
                while i < len(records):
                    batch = records[i:i + self.batch_size]
                    i += self.batch_size
                    
                    pbar.set_description(f"{Fore.GREEN}Batch {batch_num}/{total_batches}{Style.RESET_ALL}")
                    
                    xml_path = self.create_batch_xml(batch, batch_num)
                    self.send_batch_xml(xml_path, batch)
                    self.total_sent += len(batch)
                    
                    # Update progress
                    pbar.update(1)
                    
                    # Save after each batch
                    self.wb.save(self.excel_path)
                    
                    # Wait between batches
                    if i < len(records):
                        wait = random.randint(self.min_sleep, self.max_sleep)
                        pbar.set_description(f"{Fore.YELLOW}Waiting {wait}s...{Style.RESET_ALL}")
                        time.sleep(wait)
                    
                    batch_num += 1
        
        # Final summary
        end_time = datetime.now()
        duration = (end_time - start_time).total_seconds()
        
        self.print_banner("✅ COMPLETED", Fore.GREEN)
        
        summary_table = [
            ["Total Messages", self.total_sent],
            [f"{Fore.GREEN}Successful{Style.RESET_ALL}", self.total_success],
            [f"{Fore.RED}Failed{Style.RESET_ALL}", self.total_failed],
            ["Duration", f"{duration:.1f} seconds"],
            ["Excel File", str(self.excel_path.name)]
        ]
        
        print(tabulate(summary_table, tablefmt="fancy_grid"))
        print(f"\n{Fore.GREEN}{Style.BRIGHT}✓ Excel file saved successfully{Style.RESET_ALL}\n")


if __name__ == "__main__":
    # Example usage - adjust parameters as needed
    
    sender = BulkSMSSender(
        excel_path=r"C:\Users\droge\OneDrive\Desktop\sms.xlsx",
        mpe_exe_path=r"C:\Program Files (x86)\MyPhoneExplorer\MyPhoneExplorer.exe",
        mode="direct",  # "direct" for one-by-one, "batch" for XML batches
        batch_size=1,  # Only used in batch mode
        min_sleep=8,
        max_sleep=15
    )
    
    sender.run()