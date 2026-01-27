#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
KIYABO PRIMARY SCHOOL PUPIL IMPORTER v10.0 — FINAL TZ EDITION
EXACT ORIGINAL LOGIC → CLASS-BASED → ONE EXCEL FILE → ALL IN __INIT__
NO FILE CHECKS. NO MULTIPLE FILES. NO DIALOG. NO CHANGES TO LOGIC.
"""
import pyodbc
import pandas as pd
from pathlib import Path
from datetime import datetime
import re
import openpyxl
from openpyxl.styles import PatternFill
import time
from tqdm import tqdm
from rich.console import Console
from rich.table import Table
from rich import box
from rich.panel import Panel


class PrimaryPupilImporter:
    def __init__(
        self,
        excel_path: str,
        class_id: str,
        db_path: str = r"C:\Kiyabo App\backend\Kiyabo App Backend v3.0.0.accdb",
        save_folder: str = r"C:\Kiyabo App\admission",
        academic_year: int = None
    ):
        # ALL PARAMETERS IN __INIT__ — AS YOU DEMANDED
        self.excel_path = excel_path
        self.class_id = class_id.strip().upper()
        self.DB_PATH = Path(db_path)
        self.SAVE_FOLDER = Path(save_folder)
        self.SAVE_FOLDER.mkdir(parents=True, exist_ok=True)
        self.GREEN_FILL = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        self.RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        self.console = Console()
        # Set academic_year to current year if None
        self.academic_year = academic_year if academic_year is not None else datetime.now().year

    @staticmethod
    def get_class_number(class_id: str) -> int:
        # PRIMARY SCHOOL CLASS MAPPING
        mapping = {
            "BABY": 0, "MIDDLE": 1, "PRE-UNIT": 2, "GRADE 1": 3, "GRADE 2": 4, 
            "GRADE 3": 5, "GRADE 4": 6, "GRADE 5": 7, "GRADE 6": 8, "GRADE 7": 9
        }
        num = mapping.get(class_id.upper(), -1)
        if num == -1:
            raise ValueError(f"Invalid Class ID: {class_id}")
        return num

    @staticmethod
    def get_roman_class_number(class_id: str) -> str:
        """
        Convert class ID string to Roman numeral equivalent
        """
        class_mapping = {
            "BABY": "I",
            "MIDDLE": "II",
            "PRE-UNIT": "III",
            "GRADE 1": "I",
            "GRADE 2": "II",
            "GRADE 3": "III",
            "GRADE 4": "IV",
            "GRADE 5": "V",
            "GRADE 6": "VI",
            "GRADE 7": "VI"
        }
        
        return class_mapping.get(class_id.upper(), "X")

    @staticmethod
    def get_check_digit(pseudo: str) -> str:
        total = 0
        for i, ch in enumerate(reversed(pseudo)):
            n = int(ch)
            total += n if i % 2 == 0 else (n * 2 if n * 2 <= 9 else n * 2 - 9)
        return str((10 - total % 10) % 10)

    @staticmethod
    def escape_single_quote(s) -> str:
        if pd.isna(s): return ""
        return str(s).strip().replace("'", "''")

    @staticmethod
    def clean_phone(phone) -> str:
        if pd.isna(phone) or str(phone).strip() == "":
            return None
        digits = re.sub(r"\D", "", str(phone))
        cleaned_digits = digits[-9:] if len(digits) > 9 else digits
        return f"+255{cleaned_digits}" if cleaned_digits else None

    def name_exists(self, cur, first, middle, surname) -> bool:
        if not all([first, middle, surname]): return False
        sql = "SELECT COUNT(*) FROM [tbl_pupil_academic_info] WHERE first_name=? AND middle_name=? AND surname=?"
        cur.execute(sql, first, middle, surname)
        return cur.fetchone()[0] > 0

    def get_last_pupil_id_once(self, cur, class_number) -> str:
        year_part = datetime.now().year - class_number if class_number == 0 else datetime.now().year + 1 - class_number
        criteria = f"{year_part}{class_number}"
        sql = "SELECT TOP 1 pupil_id FROM [tbl_pupil_academic_info] WHERE reg_year_criteria=? ORDER BY pupil_id DESC"
        cur.execute(sql, criteria)
        row = cur.fetchone()
        return row[0] if row else "0"

    def generate_pupil_ids_in_memory(self, last_id: str, count: int, class_number: int) -> list:
        start_serial = 1 if last_id == "0" else int(last_id[5:9]) + 1
        ids = []
        for serial in range(start_serial, start_serial + count):
            year_part = datetime.now().year - class_number if class_number == 0 else datetime.now().year + 1 - class_number
            pseudo = f"{year_part}{class_number}{serial:04d}"
            ids.append(pseudo + self.get_check_digit(pseudo))
        return ids

    def run(self):
        start_time = time.time()
        CLASS_ID = self.class_id
        class_number = self.get_class_number(CLASS_ID)

        self.console.rule(f"[bold magenta]KIYABO PRIMARY IMPORTER – CLASS {CLASS_ID}[/bold magenta]", style="magenta")
        self.console.print(f"[cyan]DB:[/cyan] {self.DB_PATH.name}")
        self.console.print(f"[cyan]XLS:[/cyan] {Path(self.excel_path).name}\n")

        df = pd.read_excel(self.excel_path, header=0, dtype=str, engine='openpyxl')
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb.active

        if df.empty:
            self.console.print("[red]Excel is empty.[/red]")
            return

        # PREVIEW (unchanged)
        self.console.print(Panel("[bold yellow]EXCEL PREVIEW (first 5 rows)[/bold yellow]"))
        preview = df[["FIRST NAME", "MIDDLE NAME", "SURNAME", "SEX", "CLASS ID"]].head(5).copy()
        preview.columns = ["FIRST", "MIDDLE", "SURNAME", "SEX", "CLASS"]
        self.console.print(preview.to_string(index=False), "\n")

        # VALIDATE ROWS
        valid_rows = []
        for idx in df.index:
            r = df.loc[idx]
            required = ["FIRST NAME", "MIDDLE NAME", "SURNAME", "SEX", "CLASS ID"]
            if any(pd.isna(r.get(col)) for col in required):
                continue
            if str(r["CLASS ID"]).strip().upper() != CLASS_ID:
                self.console.print(f"[yellow]Warning: Row {idx+2} has Class '{r['CLASS ID']}' ≠ '{CLASS_ID}'[/yellow]")
                continue
            valid_rows.append((idx, r))

        if not valid_rows:
            self.console.print(f"[red]No valid pupils found for Class {CLASS_ID}.[/red]")
            return

        self.console.print(f"[bold green]Ready: {len(valid_rows)} pupils to import.[/bold green]\n")

        # DB CONNECTION
        CONN_STR = f"Driver={{Microsoft Access Driver (*.mdb, *.accdb)}};Dbq={self.DB_PATH};"
        conn = pyodbc.connect(CONN_STR, autocommit=False)
        cur = conn.cursor()

        try:
            last_id = self.get_last_pupil_id_once(cur, class_number)
            self.console.print(f"[bold cyan]Last Pupil ID in DB: {last_id}[/bold cyan]\n")

            pupil_ids = self.generate_pupil_ids_in_memory(last_id, len(valid_rows), class_number)

            # Collections for batch insert
            pupil_records = []
            admission_records = []
            family_records = []
            health_records = []        # ← NEW: only pupil_id
            enrollment_records = []    # ← NEW: pupil_id, class_id, section_id, academic_year
            family_keys = [
                "pupil_id", "date_of_birth", "father_name", "father_occupation", "father_phone", "father_phone_alternative",
                "mother_name", "mother_occupation", "mother_phone", "mother_phone_alternative",
                "gurdian_name", "gurdian_occupation", "gurdian_relationship", "gurdian_phone", "gurdian_phone_alternative",
                "parent_name", "parent_occupation", "parent_phone", "parent_phone_alternative", "parent_relationship"
            ]

            stats = {"total": len(valid_rows), "inserted": 0, "skipped": 0, "skip_reasons": {"duplicate_name": 0}}
            table = Table(title="Processing Pupils", box=box.DOUBLE, show_header=True, header_style="bold magenta")
            for col, width in [("SN",4), ("ID",12), ("NAME",28), ("SEX",5), ("CLASS",12), ("SECTION",8), ("STATUS",10)]:
                table.add_column(col, width=width)

            serial = 1

            for i, (excel_row_idx, r) in enumerate(tqdm(valid_rows, desc="Preparing data", unit="pupil")):
                pupil_id = pupil_ids[i]
                first = self.escape_single_quote(r["FIRST NAME"]).upper()
                middle = self.escape_single_quote(r["MIDDLE NAME"]).upper()
                surname = self.escape_single_quote(r["SURNAME"]).upper()
                sex = str(r["SEX"]).upper()
                inactive = -1 if str(r.get("INACTIVE", "")).strip().upper() == "YES" else 0
                section = str(r["SECTION"]) if pd.notna(r.get("SECTION")) and str(r["SECTION"]).strip() else ""
                prem_no = str(r["PREM NO"]) if pd.notna(r.get("PREM NO")) and str(r["PREM NO"]).strip() else None
                full_name = f"{first} {middle} {surname}"
                row_offset = excel_row_idx + 2

                skip_reason = None
                if self.name_exists(cur, first.replace("''", "'"), middle.replace("''", "'"), surname.replace("''", "'")):
                    skip_reason = "Duplicate Name"
                    stats["skipped"] += 1
                    stats["skip_reasons"]["duplicate_name"] += 1
                    for c in range(2, 10): ws.cell(row=row_offset, column=c).fill = self.RED_FILL
                else:
                    stats["inserted"] += 1
                    ws.cell(row=row_offset, column=21).value = pupil_id
                    for c in range(2, 10): ws.cell(row=row_offset, column=c).fill = self.GREEN_FILL

                    roman_class_number = self.get_roman_class_number(CLASS_ID)
                    section_id = roman_class_number + section[-1].upper() if section else None

                    # 1. Academic Info (section_id moved to enrollment table)
                    pupil_records.append((
                        pupil_id, first.replace("''", "'"), middle.replace("''", "'"), surname.replace("''", "'"),
                        sex, inactive, prem_no
                    ))

                    # 2. Admission
                    admn_no = str(r["ADMISSION NO"]) if pd.notna(r.get("ADMISSION NO")) else None
                    entrance_mode = self.escape_single_quote(r.get("ENTRANCE MODE", "DIRECT")).upper()
                    former_school = self.escape_single_quote(r.get("FORMER SCHOOL", "")).upper().replace("''", "'") if pd.notna(r.get("FORMER SCHOOL")) else None
                    admission_records.append((pupil_id, admn_no, entrance_mode, former_school))
                    for c in range(10, 13): ws.cell(row=row_offset, column=c).fill = self.GREEN_FILL

                    # 3. Family Info
                    dob = r.get("DATE OF BIRTH") if pd.notna(r.get("DATE OF BIRTH")) else None
                    parent_type = str(r.get("RELATIONSHIP", "")).strip().upper()
                    parent_name = self.escape_single_quote(r.get("PARENT NAME", ""))
                    occupation = self.escape_single_quote(r.get("OCCUPATION", ""))
                    phone1 = self.clean_phone(r.get("PHONE NUMBER"))
                    phone2 = self.clean_phone(r.get("ALTERNATIVE PHONE"))
                    gurdian_rel = self.escape_single_quote(r.get("IF GUARDIAN SPECIFY", ""))

                    fam = {k: None for k in family_keys}
                    fam["pupil_id"] = pupil_id
                    fam["date_of_birth"] = dob

                    if parent_type == "FATHER":
                        fam.update({"father_name": parent_name or f"{middle} {surname}", "father_occupation": occupation,
                                   "father_phone": phone1, "father_phone_alternative": phone2,
                                   "parent_name": parent_name or f"{middle} {surname}", "parent_relationship": "FATHER",
                                   "parent_occupation": occupation, "parent_phone": phone1, "parent_phone_alternative": phone2})
                    elif parent_type == "MOTHER":
                        fam.update({"mother_name": parent_name, "mother_occupation": occupation,
                                   "mother_phone": phone1, "mother_phone_alternative": phone2,
                                   "parent_name": parent_name, "parent_relationship": "MOTHER",
                                   "parent_occupation": occupation, "parent_phone": phone1, "parent_phone_alternative": phone2})
                    elif parent_type == "GUARDIAN" or (not parent_type and (phone1 or phone2)):
                        rel = gurdian_rel or "GUARDIAN"
                        fam.update({"gurdian_name": parent_name, "gurdian_occupation": occupation,
                                   "gurdian_relationship": rel, "gurdian_phone": phone1, "gurdian_phone_alternative": phone2,
                                   "parent_name": parent_name, "parent_relationship": rel,
                                   "parent_occupation": occupation, "parent_phone": phone1, "parent_phone_alternative": phone2})

                    family_records.append([fam[k] for k in family_keys])
                    for c in range(13, 20): ws.cell(row=row_offset, column=c).fill = self.GREEN_FILL

                    # 4. Health Info – only pupil_id (one-to-many)
                    health_records.append((pupil_id,))

                    # 5. Enrollment Info – pupil_id, class_id, section_id, academic_year
                    enrollment_records.append((
                        pupil_id,
                        CLASS_ID,
                        section_id if section_id else None,
                        self.academic_year
                    ))

                status = "[red]SKIPPED[/red]" if skip_reason else "[green]INSERTED[/green]"
                table.add_row(str(serial), pupil_id if not skip_reason else "—",
                            full_name[:27] + ("…" if len(full_name) > 27 else ""), sex, CLASS_ID, section or "—", status)
                serial += 1

            self.console.print(table)

            # FAST BATCH INSERTS – NOW SAFE BECAUSE RELATIONSHIPS ARE CORRECT
            batch_size = 20

            if pupil_records:
                with tqdm(total=len(pupil_records), desc="Inserting Academic Info", unit="rec") as pbar:
                    for i in range(0, len(pupil_records), batch_size):
                        cur.executemany("""INSERT INTO [tbl_pupil_academic_info]
                            (pupil_id, first_name, middle_name, surname, sex, inactive, prem_no)
                            VALUES (?, ?, ?, ?, ?, ?, ?)""", pupil_records[i:i+batch_size])
                        conn.commit()
                        pbar.update(len(pupil_records[i:i+batch_size]))

            if admission_records:
                print("\n")
                with tqdm(total=len(admission_records), desc="Inserting Admission", unit="rec") as pbar:
                    for i in range(0, len(admission_records), batch_size):
                        cur.executemany("""INSERT INTO [tbl_pupil_admission]
                            (pupil_id, admn_no, entrance_mode, former_school) VALUES (?, ?, ?, ?)""",
                            admission_records[i:i+batch_size])
                        conn.commit()
                        pbar.update(len(admission_records[i:i+batch_size]))

            if family_records:
                print("\n")
                used = [c for i, c in enumerate(family_keys) if any(rec[i] for rec in family_records)]
                sql = f"INSERT INTO [tbl_pupil_family_info] ({', '.join(f'[{c}]' for c in used)}) VALUES ({', '.join('?' * len(used))})"
                data_batches = [[r[family_keys.index(c)] for c in used] for r in family_records]
                with tqdm(total=len(data_batches), desc="Inserting Family Info", unit="rec") as pbar:
                    for i in range(0, len(data_batches), batch_size):
                        cur.executemany(sql, data_batches[i:i+batch_size])
                        conn.commit()
                        pbar.update(len(data_batches[i:i+batch_size]))

            if health_records:
                print("\n")
                with tqdm(total=len(health_records), desc="Inserting Health Records", unit="rec") as pbar:
                    for i in range(0, len(health_records), batch_size):
                        cur.executemany("INSERT INTO [tbl_pupil_health_info] (pupil_id) VALUES (?)", health_records[i:i+batch_size])
                        conn.commit()
                        pbar.update(len(health_records[i:i+batch_size]))

            if enrollment_records:
                print("\n")
                with tqdm(total=len(enrollment_records), desc="Inserting Enrollment Records", unit="rec") as pbar:
                    for i in range(0, len(enrollment_records), batch_size):
                        cur.executemany("""INSERT INTO [tbl_pupil_enrollments]
                            (pupil_id, class_id, section_id, academic_year) VALUES (?, ?, ?, ?)""",
                            enrollment_records[i:i+batch_size])
                        conn.commit()
                        pbar.update(len(enrollment_records[i:i+batch_size]))

            # SAVE EXCEL
            timestamp = datetime.now().strftime("%d%b%Y %H%M%S")
            new_file = self.SAVE_FOLDER / f"UPLOADED Form {CLASS_ID} {timestamp}.xlsx"
            wb.save(new_file)

            # SUMMARY
            elapsed = time.time() - start_time
            mins, secs = divmod(elapsed, 60)
            summary = Table(title="IMPORT SUMMARY", box=box.ROUNDED)
            summary.add_column("Metric", style="bold cyan")
            summary.add_column("Value", style="bold green")
            summary.add_row("Total Pupils", str(stats["total"]))
            summary.add_row("Inserted", str(stats["inserted"]))
            summary.add_row("Skipped", str(stats["skipped"]))
            summary.add_row(" • Duplicate Name", str(stats["skip_reasons"]["duplicate_name"]))
            summary.add_row("Academic Records", str(len(pupil_records)))
            summary.add_row("Admission Records", str(len(admission_records)))
            summary.add_row("Family Records", str(len(family_records)))
            summary.add_row("Health Records", str(len(health_records)))
            summary.add_row("Enrollment Records", str(len(enrollment_records)))
            summary.add_row("Time Taken", f"{int(mins)} min {secs:.1f} sec")
            summary.add_row("Saved File", new_file.name)

            self.console.rule("[bold green]IMPORT COMPLETED SUCCESSFULLY[/bold green]")
            self.console.print(summary)

        except Exception as e:
            conn.rollback()
            self.console.print(f"[bold red]FATAL ERROR: {e}[/bold red]")
            raise
        finally:
            conn.close()

# USAGE — ALL IN __INIT__ — ONE EXCEL FILE
if __name__ == "__main__":
    PrimaryPupilImporter(
        excel_path=r"C:\admission\Form Grade 6 Upload.xlsx",
        class_id="Grade 6"
    ).run()