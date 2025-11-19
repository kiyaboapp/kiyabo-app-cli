#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
KIYABO A-LEVEL IMPORTER v9.0 — FINAL — NOV 11 2025 — TZ TIME: 09:58 AM EAT
→ is_present = "We have at least one student in this combination"
→ Excel COMB = comb_id (PCM, PCB, HGK, etc.)
→ After import → UPDATE is_present = -1 for all used combs
→ Auto-fix subjects: Remove General Studies (31), Add Historia (30) + Academic Comm (34)
→ TQDM + COLOURFUL TABLES + SERIOUS STYLE
→ NO BULLSHIT. NO GUESSING. 1214 STUDENTS = SUCCESS
"""
import re
import time
from pathlib import Path
from datetime import datetime
import pandas as pd
import pyodbc
import openpyxl
from openpyxl.styles import PatternFill
from tqdm import tqdm
from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich import box

console = Console()

class AlevelStudentImporter:
    def __init__(
        self,
        excel_path: str,
        class_id: str,
        db_path: str = r"C:\Kiyabo App\backend\Kiyabo App Backend v2.0.0.accdb",
        save_folder: str = r"C:\Kiyabo App\admission"
    ):
        self.excel_path = Path(excel_path)
        self.class_id = class_id.strip().upper()
        self.DB_PATH = Path(db_path)
        self.SAVE_FOLDER = Path(save_folder)
        self.SAVE_FOLDER.mkdir(parents=True, exist_ok=True)

        # COLOURS
        self.GREEN = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")
        self.RED = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        self.YELLOW = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")

        # LOAD EXCEL
        console.print(f"[bold blue]Loading Excel:[/bold blue] {self.excel_path.name}")
        self.df = pd.read_excel(self.excel_path, dtype=str,engine="openpyxl").fillna("")
        self.wb = openpyxl.load_workbook(self.excel_path)
        self.ws = self.wb.active

        # DB
        CONN_STR = f"Driver={{Microsoft Access Driver (*.mdb, *.accdb)}};Dbq={self.DB_PATH};"
        self.conn = pyodbc.connect(CONN_STR, autocommit=False)
        self.cur = self.conn.cursor()

        # ALL COMBINATIONS
        self.cur.execute("SELECT comb_id, comb_name FROM tbl_student_combs")
        rows = self.cur.fetchall()
        self.all_combs = {r[0].strip().upper(): r[1].strip() for r in rows}
        self.valid_combs = set(self.all_combs.keys())

        console.print(f"[bold magenta]Database Ready — {len(self.all_combs)} combinations loaded[/bold magenta]\n")

        self.class_number = {"V": 5, "VI": 6}[self.class_id]

    @staticmethod
    def check_digit(pseudo: str) -> str:
        total = sum(int(ch) if i % 2 == 0 else (int(ch) * 2 if int(ch) * 2 <= 9 else int(ch) * 2 - 9)
                    for i, ch in enumerate(reversed(pseudo)))
        return str((10 - total % 10) % 10)

    @staticmethod
    def esc(s): return str(s).strip().replace("'", "''") if s else ""

    @staticmethod
    def phone(s): return re.sub(r"\D", "", str(s))[-9:].zfill(9)

    def name_exists(self, f, m, s):
        self.cur.execute("SELECT 1 FROM [tbl_student_academic_info] WHERE first_name=? AND middle_name=? AND surname=?", (f, m, s))
        return bool(self.cur.fetchone())

    def last_id(self):
        y = datetime.now().year + 1 - self.class_number
        c = f"{y}{self.class_number}"
        self.cur.execute("SELECT TOP 1 student_id FROM [tbl_student_academic_info] WHERE reg_year_criteria=? ORDER BY student_id DESC", (c,))
        r = self.cur.fetchone()
        return r[0] if r else "0"

    def gen_ids(self, n):
        last = self.last_id()
        start = 1 if last == "0" else int(last[5:9]) + 1
        y = datetime.now().year + 1 - self.class_number
        return [f"{y}{self.class_number}{s:04d}{self.check_digit(f'{y}{self.class_number}{s:04d}')}" for s in range(start, start + n)]

    def update_comb_subjects(self):
        console.print("\n[bold cyan]Fixing combination subjects (Remove GS, Add Historia & Comm)...[/bold cyan]")
        start = time.time()

        # BULK DELETE GENERAL STUDIES
        self.cur.execute("DELETE FROM tbl_student_comb_subjects WHERE subject_id = 31")
        del_gs = self.cur.rowcount
        self.cur.execute("UPDATE tbl_student_subjects SET is_present = 0 WHERE subject_serial = 31")
        del_pres = self.cur.rowcount
        self.conn.commit()

        # GET ALL comb_id
        self.cur.execute("SELECT comb_id FROM tbl_student_combs")
        combs = [r[0].strip() for r in self.cur.fetchall()]

        inserts = []
        for comb in tqdm(combs, desc="Checking subjects", colour="cyan", leave=False):
            for sid in [30, 34]:
                self.cur.execute("SELECT 1 FROM tbl_student_comb_subjects WHERE comb_id=? AND subject_id=?", (comb, sid))
                if not self.cur.fetchone():
                    inserts.append((comb, sid))

        if inserts:
            self.cur.executemany("INSERT INTO tbl_student_comb_subjects (comb_id, subject_id) VALUES (?, ?)", inserts)
            self.conn.commit()

        console.print(f"[bold green]Subjects fixed in {time.time()-start:.2f}s[/bold green]")
        console.print(f"   • [red]Removed General Studies:[/red] {del_gs + del_pres}")
        console.print(f"   • [green]Added Historia & Comm:[/green] {len(inserts)}")

    def run(self):
        start_all = time.time()
        console.rule(f"[bold magenta]KIYABO A-LEVEL FORM {self.class_id} IMPORTER — 2025[/bold magenta]", style="magenta")

        # PREVIEW
        preview = self.df[["FIRST NAME", "MIDDLE NAME", "SURNAME", "SEX", "CLASS ID", "COMB"]].head(5).copy()
        preview.columns = ["FIRST", "MIDDLE", "SURNAME", "SEX", "CLASS", "COMB"]
        console.print(Panel(preview.to_string(index=False), title="[bold yellow]Excel Preview[/bold yellow]"))

        # VALIDATE
        valid = []
        invalid = []
        for idx, r in self.df.iterrows():
            req = ["FIRST NAME", "MIDDLE NAME", "SURNAME", "SEX", "CLASS ID", "COMB"]
            if any(str(r.get(c, "")).strip() == "" for c in req):
                continue
            if str(r["CLASS ID"]).strip().upper() != self.class_id:
                continue
            comb = str(r["COMB"]).strip().upper()
            if comb not in self.valid_combs:
                invalid.append((idx, comb))
                continue
            valid.append((idx, r, comb, self.all_combs[comb]))

        if not valid:
            console.print("[bold red]NO VALID STUDENTS — CHECK COMB CODES[/bold red]")
            return

        console.print(f"[bold green]Valid students: {len(valid)}[/bold green]")
        if invalid:
            console.print(f"[bold red]Invalid COMB: {len(invalid)}[/bold red]")

        # COLOUR EXCEL
        for idx, _, comb in invalid:
            row = idx + 2
            for c in range(2, 10): self.ws.cell(row, c).fill = self.YELLOW
            self.ws.cell(row, 7).value = f"INVALID: {comb}"

        ids = self.gen_ids(len(valid))
        stats = {"inserted": 0, "dup": 0, "invalid": len(invalid)}
        rec_student = []
        rec_admission = []
        rec_family = []
        combs_used = set()
        family_keys = [
            "student_id", "date_of_birth", "father_name", "father_occupation", "father_phone", "father_phone_alternative",
            "mother_name", "mother_occupation", "mother_phone", "mother_phone_alternative",
            "gurdian_name", "gurdian_occupation", "gurdian_relationship", "gurdian_phone", "gurdian_phone_alternative",
            "parent_name", "parent_occupation", "parent_phone", "parent_phone_alternative", "parent_relationship"
        ]

        # PROCESS WITH TQDM
        table = Table(title="Processing Students", box=box.DOUBLE_EDGE)
        table.add_column("SN", style="cyan")
        table.add_column("ID", style="bold green")
        table.add_column("Name", style="white")
        table.add_column("COMB", style="yellow")
        table.add_column("Status", style="bold")

        for i, (idx, r, comb_id, comb_name) in enumerate(tqdm(valid, desc="Importing", colour="green")):
            sid = ids[i]
            f = self.esc(r["FIRST NAME"]).upper()
            m = self.esc(r["MIDDLE NAME"]).upper()
            s = self.esc(r["SURNAME"]).upper()
            name = f"{f} {m} {s}"
            row = idx + 2

            if self.name_exists(f.replace("''", "'"), m.replace("''", "'"), s.replace("''", "'")):
                stats["dup"] += 1
                for c in range(2, 10): self.ws.cell(row, c).fill = self.RED
                table.add_row(str(i+1), "—", name[:28], comb_id, "[red]DUPLICATE[/red]")
            else:
                stats["inserted"] += 1
                for c in range(2, 20): self.ws.cell(row, c).fill = self.GREEN
                self.ws.cell(row, 21).value = sid

                sec = r.get("SECTION", "").strip()
                sec_id = self.class_id + sec[-1].upper() if sec else None
                prem = str(r.get("PREM NO", "")) or None
                inactive = -1 if str(r.get("INACTIVE", "")).strip().upper() == "YES" else 0

                rec_student.append((
                    sid, f.replace("''", "'"), m.replace("''", "'"), s.replace("''", "'"),
                    r["SEX"].upper(), 1, inactive, sec_id, prem, comb_id
                ))

                admn = str(r.get("ADMISSION NO", "")) or None
                ent = self.esc(r.get("ENTRANCE MODE", "DIRECT"))
                former = self.esc(r.get("FORMER SCHOOL", "")) or None
                rec_admission.append((sid, admn, ent, former))

                dob = r.get("DATE OF BIRTH") if pd.notna(r.get("DATE OF BIRTH")) else None
                rel = str(r.get("RELATIONSHIP", "")).strip().upper()
                p_name = self.esc(r.get("PARENT NAME", ""))
                occ = self.esc(r.get("OCCUPATION", ""))
                ph1 = self.phone(r.get("PHONE NUMBER"))
                ph2 = self.phone(r.get("ALTERNATIVE PHONE"))
                g_rel = self.esc(r.get("IF GUARDIAN SPECIFY", ""))

                fam = {k: None for k in family_keys}
                fam["student_id"] = sid
                fam["date_of_birth"] = dob

                if rel == "FATHER":
                    fam.update({"father_name": p_name or f"{m} {s}", "father_occupation": occ,
                               "father_phone": ph1, "father_phone_alternative": ph2,
                               "parent_name": p_name or f"{m} {s}", "parent_relationship": "FATHER",
                               "parent_occupation": occ, "parent_phone": ph1, "parent_phone_alternative": ph2})
                elif rel == "MOTHER":
                    fam.update({"mother_name": p_name, "mother_occupation": occ,
                               "mother_phone": ph1, "mother_phone_alternative": ph2,
                               "parent_name": p_name, "parent_relationship": "MOTHER",
                               "parent_occupation": occ, "parent_phone": ph1, "parent_phone_alternative": ph2})
                else:
                    gr = g_rel or "GUARDIAN"
                    fam.update({"gurdian_name": p_name, "gurdian_occupation": occ,
                               "gurdian_relationship": gr, "gurdian_phone": ph1, "gurdian_phone_alternative": ph2,
                               "parent_name": p_name, "parent_relationship": gr,
                               "parent_occupation": occ, "parent_phone": ph1, "parent_phone_alternative": ph2})

                rec_family.append([fam[k] for k in family_keys])
                combs_used.add(comb_id)
                table.add_row(str(i+1), sid, name[:28], comb_id, "[green]INSERTED[/green]")

        console.print(table)

        # INSERT
        with tqdm(total=4, desc="Saving to DB", colour="magenta") as p:
            if rec_student:
                self.cur.executemany("""INSERT INTO [tbl_student_academic_info]
                    (student_id, first_name, middle_name, surname, sex, is_boarding, inactive, section_id, prem_no, comb_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""", rec_student)
                self.conn.commit()
                p.update()

            if rec_admission:
                self.cur.executemany("INSERT INTO [tbl_student_admission] (student_id, admn_no, entrance_mode, former_school) VALUES (?, ?, ?, ?)", rec_admission)
                self.conn.commit()
                p.update()

            if rec_family:
                used = [c for i,c in enumerate(family_keys) if any(r[i] for r in rec_family)]
                cols = ", ".join(f"[{c}]" for c in used)
                ph = ", ".join("?" * len(used))
                self.cur.executemany(f"INSERT INTO [tbl_student_family_info] ({cols}) VALUES ({ph})",
                                   [[r[family_keys.index(c)] for c in used] for r in rec_family])
                self.conn.commit()
                p.update()

            if combs_used:
                ph = ",".join("?" for _ in combs_used)
                self.cur.execute(f"UPDATE tbl_student_combs SET is_present = -1 WHERE comb_id IN ({ph})", list(combs_used))
                self.conn.commit()
                p.update()

        # SAVE EXCEL
        ts = datetime.now().strftime("%d%b%Y_%H%M%S")
        saved = self.SAVE_FOLDER / f"UPLOADED_ALEVEL_{self.class_id}_{ts}.xlsx"
        self.wb.save(saved)

        # SUBJECT FIX
        self.update_comb_subjects()

        # FINAL
        elapsed = time.time() - start_all
        mins, secs = divmod(elapsed, 60)
        summary = Table(title="KIYABO A-LEVEL IMPORT COMPLETE", box=box.ROUNDED)
        summary.add_column("Item", style="bold cyan")
        summary.add_column("Value", style="bold green")
        for k, v in [
            ("Valid Rows", str(len(valid))),
            ("Inserted", str(stats["inserted"])),
            ("Duplicates", str(stats["dup"])),
            ("Invalid COMB", str(stats["invalid"])),
            ("Combs Now Present", str(len(combs_used))),
            ("Time", f"{int(mins)}m {secs:.1f}s"),
            ("Saved", saved.name)
        ]:
            summary.add_row(k, v)

        console.rule("[bold green]KIYABO A-LEVEL 2025 — SUCCESS[/bold green]")
        console.print(summary)
        self.conn.close()


if __name__ == "__main__":
    importer = AlevelStudentImporter(
        excel_path=r"C:\admission\A-Level Form V 2025.xlsx",
        class_id="V"
    )
    importer.run()